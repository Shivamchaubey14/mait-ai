"""
The assignment workbook the portal hands out (SRS §6.2.2).

Handed over already filled with the current mapping rather than as an empty form. An admin
maintaining which Mait covers which MPP is almost always changing a handful of rows in a list
of three thousand, and a blank template makes them retype the three thousand they are not
changing — which is where the errors come from.

The headers here are the ones ``columns.ASSIGNMENT`` reads back, so a file that goes out and
comes back unedited is a no-op rather than a validation failure.
"""

from __future__ import annotations

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import MPP

# "Mait", not "Sahayak". The MPP master has a Sahayak column too, and it means the person who
# staffs the collection point — a different job from the AI technician who covers it. Naming
# both the same is what put 3,110 villagers in the Mait roster.
HEADERS = [
    ("MPP Code", 14),
    ("MPP Name", 30),
    ("District", 12),
    ("Mait Vendor", 18),
    ("Mait Name", 26),
    ("Mobile No", 14),
]

# The portal's ink and green, so the file looks like it came from the same product.
HEADER_FILL = PatternFill("solid", fgColor="253D4E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
# Unassigned rows tinted, because they are the ones the sheet was downloaded to fix.
BLANK_FILL = PatternFill("solid", fgColor="FFF8E9")


def build_assignment_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Assignments"

    for index, (title, width) in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width

    # Unassigned first: they are the work. Then by name, which is how an operator scans for
    # the one they came to change.
    rows = MPP.objects.select_related("mait").order_by("mait__name", "mpp_name")
    unassigned = [mpp for mpp in rows if mpp.mait_id is None]
    assigned = [mpp for mpp in rows if mpp.mait_id is not None]

    line = 2
    for mpp in [*unassigned, *assigned]:
        values = [
            mpp.mpp_code,
            mpp.mpp_name,
            mpp.district_code,
            mpp.mait.sahayak_vendor_code if mpp.mait else "",
            mpp.mait.name if mpp.mait else "",
            mpp.mait.mobile_no if mpp.mait else "",
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=line, column=index, value=value)
            # Codes are text. Left as numbers, Excel drops the leading zero on "001302" and
            # the file comes back naming an MPP that does not exist.
            if index in (1, 4, 6):
                cell.number_format = "@"
            if mpp.mait_id is None:
                cell.fill = BLANK_FILL
        line += 1

    sheet.freeze_panes = "A2"
    return workbook


def assignment_template_response() -> HttpResponse:
    workbook = build_assignment_workbook()
    stamp = timezone.localdate().isoformat()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="mait-mpp-assignments-{stamp}.xlsx"'
    workbook.save(response)
    workbook.close()
    return response
