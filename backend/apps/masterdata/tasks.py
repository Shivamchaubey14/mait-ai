"""
SAP master-data import (SRS §6.1).

The Member Master is ~105,000 rows across 54 columns in a ~28 MB workbook, so this runs as a
Celery job with progress reporting and never inline on a request (SRS §6.1.6).

Four behaviours matter and are easy to get wrong:

* **Header detection.** ``Member.xlsx`` opens with four title rows and a blank line; the real
  header is row 6. The first row cannot be assumed (SRS §6.1.2).
* **Upsert by natural key**, so re-uploading last month's file refreshes rather than
  duplicates (SRS §6.1.3).
* **Partial success.** Invalid rows are skipped and reported; valid rows still commit
  (SRS §6.1.4). Rejecting 105k rows because twelve have a malformed mobile number would make
  the feature unusable.
* **Streamed reading.** ``read_only=True`` keeps openpyxl from materialising the whole
  workbook, which is the difference between a few hundred MB of memory and a killed worker.

Column spellings live in ``columns.py`` — see the notes there on the duplicate ``Mobile No``
in ``Sahyak.xlsx`` and SAP's *Tahsil* spelling.
"""

from __future__ import annotations

import datetime as dt
import logging
from collections.abc import Iterator
from typing import Any

from celery import shared_task
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from . import columns as cols
from .models import MAX_ERRORS_STORED, MPP, DataUploadLog, Mait, Member

logger = logging.getLogger(__name__)

CHUNK_SIZE = 1000
# Rows between progress writes. Smaller than a chunk on purpose: the chunk size is a
# transaction decision and one write per thousand rows is not a progress bar.
PROGRESS_EVERY = 100
HEADER_SEARCH_ROWS = 15  # Member.xlsx puts its header on row 6

REQUIRED_COLUMNS: dict[str, tuple[str, ...]] = {
    DataUploadLog.UploadType.MEMBER: cols.REQUIRED_MEMBER,
    DataUploadLog.UploadType.MPP: cols.REQUIRED_MPP,
    DataUploadLog.UploadType.MAIT: cols.REQUIRED_VENDOR,
    DataUploadLog.UploadType.ASSIGNMENT: cols.REQUIRED_ASSIGNMENT,
}


@shared_task(bind=True, name="apps.masterdata.tasks.process_master_upload")
def process_master_upload(self, upload_id: int) -> dict[str, int]:
    """Parse, validate and upsert one uploaded SAP workbook."""
    upload = DataUploadLog.objects.get(pk=upload_id)
    upload.status = DataUploadLog.Status.PROCESSING
    upload.celery_task_id = self.request.id or ""
    upload.started_at = timezone.now()
    upload.save(update_fields=["status", "celery_task_id", "started_at", "updated_at"])

    try:
        result = _import_workbook(upload)
    except Exception as exc:
        logger.exception("Master upload %s failed", upload_id)
        upload.status = DataUploadLog.Status.FAILED
        upload.error_report = [{"row": None, "error": str(exc)[:500]}]
        upload.finished_at = timezone.now()
        upload.save(update_fields=["status", "error_report", "finished_at", "updated_at"])
        raise

    upload.status = (
        DataUploadLog.Status.COMPLETED_WITH_ERRORS
        if upload.failed_rows
        else DataUploadLog.Status.COMPLETED
    )
    upload.finished_at = timezone.now()
    upload.save(update_fields=["status", "finished_at", "updated_at"])
    return result


class ImportContext:
    """
    Per-run state shared across rows.

    Exists for two reasons. The MPP lookup is cached because resolving it per row costs one
    query per member, and there are 105,000 of them against a table of only ~3,100 — the
    whole map fits in memory many times over.

    ``seen_keys`` catches a natural key appearing twice *within one file*. The real Member
    export contains 50 such duplicates, and because rows are upserted, the later row would
    silently overwrite the earlier one and the count would look clean. Flagging them turns a
    silent data loss into a reported row (SRS §6.1.4).
    """

    def __init__(self, upload_type: str = "") -> None:
        self.mpp_ids: dict[str, int] = {}
        self.seen_keys: set[str] = set()
        # Carried so a rejected row can be reported with the cells that identify it, which
        # differ per file — see `columns.IDENTITY`.
        self.upload_type = upload_type

    def load_mpp_map(self) -> None:
        self.mpp_ids = dict(MPP.objects.values_list("mpp_code", "id"))

    def check_duplicate(self, key: str) -> None:
        if key in self.seen_keys:
            raise ValueError(f"Duplicate key '{key}' appears more than once in this file.")
        self.seen_keys.add(key)


def _import_workbook(upload: DataUploadLog) -> dict[str, int]:
    handler = {
        DataUploadLog.UploadType.MEMBER: _upsert_member,
        DataUploadLog.UploadType.MAIT: _upsert_vendor,
        DataUploadLog.UploadType.MPP: _upsert_mpp_and_sahayak,
        DataUploadLog.UploadType.ASSIGNMENT: _upsert_assignment,
    }[upload.upload_type]

    context = ImportContext(upload.upload_type)
    if upload.upload_type == DataUploadLog.UploadType.MEMBER:
        context.load_mpp_map()

    workbook = load_workbook(upload.file, read_only=True, data_only=True)
    try:
        return _read_and_apply(upload, workbook, handler, context)
    finally:
        # In a `finally` because a read-only workbook holds the file open, and openpyxl's
        # read-only mode leaks the handle if the run does not reach its own close(). On Windows
        # that handle then blocks anything touching the stored file — including deleting it.
        workbook.close()


def _read_and_apply(upload: DataUploadLog, workbook, handler, context) -> dict[str, int]:
    sheet = workbook.active

    header_row_index, headers = _detect_header(sheet, upload.upload_type)
    missing = _missing_columns(REQUIRED_COLUMNS[upload.upload_type], headers)
    if missing:
        # SRS §6.1.2 — reject the whole file, and say what was actually found so the fix is
        # obvious rather than a guessing game.
        found = ", ".join(sorted(h for h in headers if h)[:25])
        raise ValueError(f"Required column(s) missing: {', '.join(missing)}. Found: {found}")

    # The bar has to have something to be a fraction of. `progress_percent` is
    # processed/total_rows and total_rows was only written once the import had finished, so
    # every upload — including the 105k-row Member master — reported 0% for its entire run and
    # then jumped to 100% when there was nothing left to wait for. The sheet's own declared
    # dimension is an estimate, which is all a bar needs; the exact count is written at the end
    # as before. Some writers omit the dimension, and then there is simply no estimate to give.
    declared_rows = max(0, (sheet.max_row or 0) - header_row_index)
    if declared_rows:
        DataUploadLog.objects.filter(pk=upload.pk).update(
            total_rows=declared_rows, updated_at=timezone.now()
        )

    errors: list[dict[str, Any]] = []
    success = failed = processed = 0
    batch: list[dict[str, Any]] = []

    for row_number, row in _iter_rows(sheet, header_row_index, headers):
        processed += 1
        batch.append({"row_number": row_number, "data": row})

        if len(batch) >= CHUNK_SIZE:
            ok, bad = _commit_batch(batch, handler, errors, context)
            success += ok
            failed += bad
            batch = []
            _report_progress(upload, processed, success, failed)
        elif processed % PROGRESS_EVERY == 0:
            # Reading runs ahead of applying by up to one chunk, which is exactly what the two
            # stage tiles say it does. Without this, a file of a thousand rows commits once —
            # at the end — and the bar has nothing to report until there is nothing left to
            # wait for, which is the state the operator described as "stuck at 0%".
            _report_progress(upload, processed, success, failed)

    if batch:
        ok, bad = _commit_batch(batch, handler, errors, context)
        success += ok
        failed += bad

    upload.total_rows = processed
    upload.processed_rows = processed
    upload.success_rows = success
    upload.failed_rows = failed
    upload.error_report = errors[:MAX_ERRORS_STORED]
    upload.save(
        update_fields=[
            "total_rows",
            "processed_rows",
            "success_rows",
            "failed_rows",
            "error_report",
            "updated_at",
        ]
    )
    logger.info("Upload %s finished: %s ok, %s failed of %s", upload.id, success, failed, processed)
    return {"total": processed, "success": success, "failed": failed}


def _missing_columns(required, headers) -> list[str]:
    """
    Which required columns the sheet does not have.

    An entry may be a nested tuple, meaning any one of those spellings satisfies it: the same
    export has shipped with different header sets, and demanding every alias would reject both
    versions of a file that is perfectly readable.
    """
    present = set(headers)
    missing = []
    for entry in required:
        options = (entry,) if isinstance(entry, str) else entry
        if not present.intersection(options):
            missing.append(" or ".join(options))
    return sorted(missing)


def _commit_batch(batch, handler, errors, context: ImportContext) -> tuple[int, int]:
    """
    Commit one chunk.

    Each row gets its own savepoint so a single bad row cannot roll back the other 999.
    That is what makes partial success work (SRS §6.1.4).
    """
    ok = bad = 0
    for item in batch:
        try:
            with transaction.atomic():
                handler(item["data"], context)
            ok += 1
        except Exception as exc:
            bad += 1
            if len(errors) < MAX_ERRORS_STORED:
                errors.append(
                    {
                        "row": item["row_number"],
                        "error": str(exc)[:300],
                        # Read off the row that failed, not looked up afterwards: the record
                        # was rejected, so there is nothing in the database to look up. This is
                        # the only place the operator's own values still exist.
                        "fields": cols.identity_of(item["data"], context.upload_type),
                    }
                )
    return ok, bad


def _report_progress(upload: DataUploadLog, processed: int, success: int, failed: int) -> None:
    """Update the counters the progress endpoint polls (SRS §6.1.6)."""
    DataUploadLog.objects.filter(pk=upload.pk).update(
        processed_rows=processed,
        success_rows=success,
        failed_rows=failed,
        updated_at=timezone.now(),
    )


def _detect_header(sheet, upload_type: str) -> tuple[int, list[str]]:
    """
    Find the header row (SRS §6.1.2).

    ``Member.xlsx`` carries a company name, a description, a status line, a member count and
    a blank row before the real header on row 6, so the first row cannot be assumed.

    The header is the first row within the search window containing every required column —
    where a required entry may be a nested tuple of acceptable spellings, so the same export
    is found under either of its header sets.
    """
    required = REQUIRED_COLUMNS[upload_type]
    for index, raw in enumerate(
        sheet.iter_rows(min_row=1, max_row=HEADER_SEARCH_ROWS, values_only=True)
    ):
        headers = cols.build_header_index(raw)
        if not _missing_columns(required, headers):
            return index + 1, headers
    expected = ", ".join(
        entry if isinstance(entry, str) else " or ".join(entry) for entry in required
    )
    raise ValueError(
        f"Could not locate a header row in the first {HEADER_SEARCH_ROWS} rows. "
        f"Expected column(s): {expected}."
    )


def _iter_rows(sheet, header_row_index: int, headers: list[str]) -> Iterator[tuple[int, dict]]:
    """Yield (row_number, {normalised_column: value}) for each non-empty row below the header."""
    for offset, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True)):
        if row is None or all(cell in (None, "") for cell in row):
            continue
        yield header_row_index + 1 + offset, dict(zip(headers, row, strict=False))


# --------------------------------------------------------------------------------------
# Value coercion
# --------------------------------------------------------------------------------------


def _clean(value) -> str:
    if value in (None, "", "None"):
        return ""
    text = str(value).strip()
    # openpyxl hands numeric-looking codes back as floats; "001301.0" is not an MPP code.
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _mobile(value) -> str:
    """
    Normalise an Indian mobile number.

    SAP exports carry these inconsistently — as floats, with +91, with spaces. Returns ""
    when the value cannot be salvaged rather than guessing, because a wrong number means the
    payment authorisation OTP goes to a stranger (SRS §6.5).
    """
    raw = _clean(value).replace(" ", "").replace("-", "")
    if raw.startswith("+91"):
        raw = raw[3:]
    elif raw.startswith("91") and len(raw) == 12:
        raw = raw[2:]
    return raw if len(raw) == 10 and raw[0] in "6789" and raw.isdigit() else ""


def _int_or_none(value) -> int | None:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _date_or_none(value):
    """
    Parse a SAP date.

    9999-12-31 is SAP's "no end date" sentinel and is stored as NULL — a real date that far
    out would distort any range query it landed in.
    """
    if value in (None, "", "None"):
        return None
    if isinstance(value, dt.datetime):
        parsed = value.date()
    elif isinstance(value, dt.date):
        parsed = value
    else:
        text = str(value).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y%m%d"):
            try:
                parsed = dt.datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue
        else:
            return None
    return None if parsed.year >= 9999 else parsed


def _is_active(value) -> bool:
    """SAP marks an active row with 'X'; the Member export uses 'Yes'."""
    return _clean(value).lower() in ("x", "yes", "true", "1", "active", "y")


# --------------------------------------------------------------------------------------
# Per-type upserts
# --------------------------------------------------------------------------------------


def _upsert_mpp_and_sahayak(row: dict, context: ImportContext) -> None:
    """
    Upsert one row of ``Sahyak.xlsx``.

    **This file is MPP data. It does not define Maits.**

    It used to. Each row carries an MPP and the Sahayak who staffs it, and the Sahayak column
    was turned into a ``Mait`` record — which produced one pseudo-Mait per village, 3,110 of
    them, each "covering" exactly one MPP. They are not Maits. A Sahayak runs a single
    collection point and takes the milk in; a Mait is the AI technician who covers many of
    them, and the real roster is the ZMAI vendor export (SRS §18.2, now settled).

    So the Sahayak is kept here as what it is — the contact for this collection point — and
    ``mait`` is deliberately absent from the defaults below. Coverage is assigned from the
    assignment sheet (SRS §6.2.2), and a master refresh must not silently undo it.
    """
    mpp_code = _clean(cols.pick(row, *cols.MPP["mpp_code"]))
    if not mpp_code:
        raise ValueError("MPP Code is blank.")

    MPP.objects.update_or_create(
        mpp_code=mpp_code,
        defaults={
            "plant_code": _clean(cols.pick(row, *cols.MPP["plant_code"])),
            "plant_name": _clean(cols.pick(row, *cols.MPP["plant_name"])),
            "mpp_name": _clean(cols.pick(row, *cols.MPP["mpp_name"])),
            "mpp_category": _clean(cols.pick(row, *cols.MPP["mpp_category"])),
            "mpp_sub_category": _clean(cols.pick(row, *cols.MPP["mpp_sub_category"])),
            "state_code": _clean(cols.pick(row, *cols.MPP["state_code"])),
            "district_code": _clean(cols.pick(row, *cols.MPP["district_code"])),
            "tehsil_code": _clean(cols.pick(row, *cols.MPP["tehsil_code"])),
            "panchayat_code": _clean(cols.pick(row, *cols.MPP["panchayat_code"])),
            "village_code": _clean(cols.pick(row, *cols.MPP["village_code"])),
            "hamlet_code": _clean(cols.pick(row, *cols.MPP["hamlet_code"])),
            "mobile_no": _mobile(cols.pick(row, *cols.MPP["mobile_no"])),
            "address_line": _clean(cols.pick(row, *cols.MPP["address_line"])),
            "is_active": _is_active(cols.pick(row, *cols.MPP["is_active"])),
            "start_date": _date_or_none(cols.pick(row, *cols.MPP["start_date"])),
            "end_date": _date_or_none(cols.pick(row, *cols.MPP["end_date"])),
            "revival_date": _date_or_none(cols.pick(row, *cols.MPP["revival_date"])),
            # The person at the collection point, recorded as a contact and nothing more.
            "sahayak_vendor_code": _clean(cols.pick(row, *cols.SAHAYAK["sahayak_vendor_code"])),
            "sahayak_name": _clean(cols.pick(row, *cols.SAHAYAK["name"])),
            "sahayak_mobile_no": _mobile(cols.pick(row, *cols.SAHAYAK["mobile_no"])),
        },
    )


def _upsert_vendor(row: dict, context: ImportContext) -> None:
    """
    Upsert one row of ``Maits Vendor C.xlsx``.

    This file's vendor key occupies a different number range (9900000000+) from the ``Sahayak
    Vendor`` codes in ``Sahyak.xlsx`` (5500000003+), so rows loaded here link to no MPP —
    coverage is assigned separately, from the assignment sheet (SRS §6.2.2).

    **Blank does not mean "clear it."** The export has shipped with columns renamed and, in
    the EXPORT_* variant, with the PAN column carrying no header at all. Overwriting every
    field unconditionally would silently wipe the PAN, GST and bank details of every Mait the
    moment a slightly different export was uploaded, and nothing on any screen would say so.
    So only the fields the row actually carries are written.
    """
    vendor_code = _clean(cols.pick(row, *cols.VENDOR["sahayak_vendor_code"]))
    if not vendor_code:
        raise ValueError("Vendor code is blank.")

    incoming = {
        "name": _clean(cols.pick(row, *cols.VENDOR["name"])),
        "mobile_no": _mobile(cols.pick(row, *cols.VENDOR["mobile_no"])),
        "pan_no": _clean(cols.pick(row, *cols.VENDOR["pan_no"])),
        "aadhar_no": _clean(cols.pick(row, *cols.VENDOR["aadhar_no"])),
        "gst_no": _clean(cols.pick(row, *cols.VENDOR["gst_no"]))[:20],
        "bank_account_no": _clean(cols.pick(row, *cols.VENDOR["bank_account_no"])),
        "ifsc_code": _clean(cols.pick(row, *cols.VENDOR["ifsc_code"]))[:15],
    }
    supplied = {field: value for field, value in incoming.items() if value}

    mait = Mait.objects.filter(sahayak_vendor_code=vendor_code).first()
    if mait is None:
        Mait.objects.create(sahayak_vendor_code=vendor_code, is_active=True, **supplied)
        return

    changed = [field for field, value in supplied.items() if getattr(mait, field) != value]
    if not mait.is_active:
        mait.is_active = True
        changed.append("is_active")
    if changed:
        for field in changed:
            if field != "is_active":
                setattr(mait, field, supplied[field])
        mait.save(update_fields=[*changed, "updated_at"])


def _upsert_member(row: dict, context: ImportContext) -> None:
    """Upsert one row of ``Member.xlsx``."""
    member_code = _clean(cols.pick(row, *cols.MEMBER["member_code"]))
    if not member_code:
        raise ValueError("Member code is blank.")

    context.check_duplicate(member_code)

    mpp_code = _clean(cols.pick(row, *cols.MEMBER["mpp_code"]))
    mpp_id = context.mpp_ids.get(mpp_code)
    if mpp_id is None:
        # Skipped rather than auto-created: inventing an MPP from a member row would place a
        # member at a collection point that may not exist (SRS §6.1.4).
        raise ValueError(f"MPP '{mpp_code}' not found. Upload the MPP master first.")

    Member.objects.update_or_create(
        member_code=member_code,
        defaults={
            "mpp_id": mpp_id,
            "member_name": _clean(cols.pick(row, *cols.MEMBER["member_name"])),
            "father_husband_name": _clean(cols.pick(row, *cols.MEMBER["father_husband_name"])),
            "gender": _clean(cols.pick(row, *cols.MEMBER["gender"]))[:10],
            "age": _int_or_none(cols.pick(row, *cols.MEMBER["age"])),
            "category": _clean(cols.pick(row, *cols.MEMBER["category"]))[:30],
            "education": _clean(cols.pick(row, *cols.MEMBER["education"]))[:50],
            "social_class": _clean(cols.pick(row, *cols.MEMBER["social_class"]))[:30],
            "sap_vendor_code": _clean(cols.pick(row, *cols.MEMBER["sap_vendor_code"]))[:20],
            "form_no": _clean(cols.pick(row, *cols.MEMBER["form_no"]))[:20],
            "folio_no": _clean(cols.pick(row, *cols.MEMBER["folio_no"]))[:20],
            "mobile_no": _mobile(cols.pick(row, *cols.MEMBER["mobile_no"])),
            "aadhar_no": _clean(cols.pick(row, *cols.MEMBER["aadhar_no"])),
            "cattle_holding": _int_or_none(cols.pick(row, *cols.MEMBER["cattle_holding"])),
            "bank_ac_no": _clean(cols.pick(row, *cols.MEMBER["bank_ac_no"])),
            "bank_name": _clean(cols.pick(row, *cols.MEMBER["bank_name"]))[:100],
            "bank_branch": _clean(cols.pick(row, *cols.MEMBER["bank_branch"]))[:100],
            "ifsc_code": _clean(cols.pick(row, *cols.MEMBER["ifsc_code"]))[:15],
            "activation_status": _clean(cols.pick(row, *cols.MEMBER["activation_status"]))[:20],
            "activation_date": _date_or_none(cols.pick(row, *cols.MEMBER["activation_date"])),
            "deactivation_date": _date_or_none(cols.pick(row, *cols.MEMBER["deactivation_date"])),
            "remarks": _clean(cols.pick(row, *cols.MEMBER["remarks"])),
        },
    )


def _upsert_assignment(row: dict, context: ImportContext) -> None:
    """
    Upsert one row of the assignment workbook.

    This file says one thing: which Mait covers which MPP. An MPP is the village collection
    point a Mait works — it is the area marker the whole app scopes on, so getting a row
    wrong moves a Mait's members, their animals and their permission to record an AI event
    (SRS §6.2.2–6.2.3).

    That is why nothing here creates an MPP. MPPs come from SAP, and a typo in a code must
    read as "no such MPP" rather than quietly bringing a new one into existence for a Mait to
    stare at. A Mait, by contrast, may be created — a genuinely new Sahayak has to be able to
    start somewhere — but only when the row names them, because a bare vendor code with no
    name produces a nameless row nobody can identify afterwards.

    A blank Sahayak column is not a missing value. It is the instruction to unassign, which
    is how an MPP is taken off a Mait who has left.
    """
    mpp_code = _clean(cols.pick(row, *cols.ASSIGNMENT["mpp_code"]))
    if not mpp_code:
        raise ValueError("MPP Code is blank.")

    # One MPP has one Mait, so the same MPP twice in a file is a contradiction rather than a
    # repeat: the later row would silently win and the count would look clean.
    context.check_duplicate(mpp_code)

    try:
        mpp = MPP.objects.get(mpp_code=mpp_code)
    except MPP.DoesNotExist:
        raise ValueError(
            f"No MPP with code '{mpp_code}'. MPPs come from the SAP master — check the code, "
            "or upload the MPP master first."
        ) from None

    vendor_code = _clean(cols.pick(row, *cols.ASSIGNMENT["sahayak_vendor_code"]))
    if not vendor_code:
        mpp.mait = None
        mpp.save(update_fields=["mait", "updated_at"])
        return

    name = _clean(cols.pick(row, *cols.ASSIGNMENT["name"]))
    mobile = _mobile(cols.pick(row, *cols.ASSIGNMENT["mobile_no"]))
    raw_mobile = _clean(cols.pick(row, *cols.ASSIGNMENT["mobile_no"]))

    # Salvaged or refused, never guessed: this number is the Mait's only way into the app and
    # the channel their OTP goes to, so a value we could not parse is an error rather than a
    # silent blank (SRS §6.5).
    if raw_mobile and not mobile:
        raise ValueError(f"'{raw_mobile}' is not a usable Indian mobile number.")

    mait = Mait.objects.filter(sahayak_vendor_code=vendor_code).first()
    if mait is not None and not mait.is_active:
        # Almost always a Sahayak code pasted in from the MPP master. Those are the people who
        # staff a collection point, not the Maits who cover it, and they were retired for
        # exactly that reason — assigning to one gives the MPP to nobody who can work it.
        raise ValueError(
            f"'{vendor_code}' is a retired record ({mait.name}) and cannot be given an MPP. "
            "Use a code from the current Mait roster."
        )
    if mait is None:
        if not name:
            raise ValueError(
                f"Vendor code '{vendor_code}' is not on record, and the row gives no name "
                "to create them with."
            )
        mait = Mait.objects.create(
            sahayak_vendor_code=vendor_code,
            name=name,
            mobile_no=mobile,
            is_active=True,
        )
    else:
        # Only what the row actually carries. A blank name or mobile leaves the existing one
        # alone — the file is an assignment sheet, not a replacement for the Sahayak master,
        # and clearing a mobile would lock a working Mait out of the app.
        changed = []
        if name and mait.name != name:
            mait.name = name
            changed.append("name")
        if mobile and mait.mobile_no != mobile:
            mait.mobile_no = mobile
            changed.append("mobile_no")
        if changed:
            mait.save(update_fields=[*changed, "updated_at"])

    mpp.mait = mait
    mpp.save(update_fields=["mait", "updated_at"])
