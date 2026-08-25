"""
Handing a master back the way it came in — as a workbook nobody can quietly edit.

An admin's question here is not "what is the template". The templates are the SAP exports
themselves, and this portal has always said so. The question is **what did we last load**:
before re-uploading a corrected Member master somebody wants to open the one currently in
force, check a column, and be sure they are looking at what the platform is actually running
on. Until now the only copy was on whichever laptop uploaded it.

So this rebuilds the stored upload as a fresh workbook and locks it.

**"Locked" is Excel's own sheet protection, and that is worth being exact about.** It stops the
accident — a stray keystroke in a cell, a column dragged, a row deleted on the way to
somewhere else — which is the whole risk here, because a master that reaches somebody subtly
altered is worse than no file at all. It is *not* encryption and it is not a permission: the
protection can be removed by anyone who means to, and openpyxl itself will do it. Nothing on
the screen claims otherwise, and nothing downstream should treat one of these as evidence.

**Rebuilt rather than served.** The stored file could simply be streamed back, and it would be
faster — but it would arrive editable, and re-emitting is also what lets the sheet carry a
provenance banner saying which upload this is and when it landed. A copy of a master with no
date on it is the thing that gets mistaken for the current one three weeks later.

**Formatted on the way out**, because a master is opened to be read rather than to be tidied.
Columns are sized to what is actually in them, the header row is frozen so it survives a scroll
through a hundred thousand rows, and the filter arrows are left usable even though every cell is
locked — sorting a column is reading, not editing. A file that opens as a wall of `#####` over a
header that scrolls away is one an admin fixes by hand before they can use it, every time.

The formatting has to be decided before a single row is written: a write-only sheet takes its
widths and its frozen pane *before* the first append, and set afterwards they are silently
discarded — the freeze vanishes and the widths come back as numbers openpyxl chose. That is
exactly how this file was written the first time, and it looked right until somebody opened the
workbook.

The first attempt at fixing it opened the source twice, once to measure and once to copy, which
took a 482 KB master from three seconds to five: every open re-parses the shared-string table,
and on the Member master that is the bulk of the work. So the opening rows are pulled into
memory instead — a few hundred of them, bounded — measured, and then written out ahead of the
rest of the stream. One open, one pass.

Streamed at both ends: `read_only` on the way in and `write_only` on the way out, because the
Member master is ~28 MB and about a hundred thousand rows, and materialising that twice is a
worker held for the length of a download.
"""

from __future__ import annotations

import io
from itertools import chain, islice

from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import DataUploadLog

#: Beyond this the honest answer is the database, not a spreadsheet. The masters sit well
#: inside it; the cap is here so one malformed upload cannot turn a download into an outage.
MAX_SNAPSHOT_ROWS = 200_000

#: How many rows are held in memory to size the columns before writing begins. Bounded, so the
#: cost is a few hundred rows however large the master is. The masters are SAP exports where a
#: column's width is settled by its first screenful — a code six characters wide on row 200 is
#: six wide on row 100,000.
WIDTH_SAMPLE_ROWS = 400

#: How often the builder reports its position while copying rows. See `build_snapshot`.
PROGRESS_EVERY_ROWS = 250

#: Bounds on a measured column. The floor keeps a two-letter header from collapsing to nothing;
#: the ceiling stops one long address column from pushing every other column off the screen,
#: which is the failure mode of naive autofit.
MIN_COLUMN_WIDTH = 9
MAX_COLUMN_WIDTH = 46

#: Sheet protection needs *a* password, or Excel offers to unprotect with a blank one and the
#: friction is zero. It is deliberately not a secret — see the module docstring. Keeping it in
#: the open is more honest than implying the file is sealed.
SHEET_PASSWORD = "mait-ai"  # noqa: S105 - not a credential; see above.

#: The portal's own ink header band, so the file looks like it came from here.
_HEADER_FILL = PatternFill("solid", fgColor="FF253D4E")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_TITLE_FONT = Font(bold=True, size=12, color="FF253D4E")
_NOTE_FONT = Font(italic=True, size=9, color="FF7A8893")


def latest_upload(upload_type: str) -> DataUploadLog | None:
    """
    The most recent upload of this master that actually landed.

    Not simply the most recent row. A queued or failed upload is a file that never became the
    master, and handing one back as "what we are running on" would be a lie of exactly the
    kind this feature exists to prevent. `completed_with_errors` counts: the rows that passed
    are in force, and the rejected ones already have their own report.
    """
    return (
        DataUploadLog.objects.filter(
            upload_type=upload_type,
            status__in=[
                DataUploadLog.Status.COMPLETED,
                DataUploadLog.Status.COMPLETED_WITH_ERRORS,
            ],
        )
        .select_related("uploaded_by")
        .order_by("-created_at")
        .first()
    )


def _provenance(upload: DataUploadLog) -> list[tuple[str, Font]]:
    """Where this file came from, written on the file itself."""
    landed = timezone.localtime(upload.finished_at or upload.created_at)
    who = upload.uploaded_by.full_name or upload.uploaded_by.username
    return [
        (
            f"{upload.get_upload_type_display()} — as loaded on "
            f"{landed.strftime('%d %b %Y at %H:%M')}",
            _TITLE_FONT,
        ),
        (
            f"Source file: {upload.file_name} · {upload.success_rows:,} of "
            f"{upload.total_rows:,} rows accepted · uploaded by {who}",
            _NOTE_FONT,
        ),
        (
            "Read-only copy. Editing here changes nothing — correct the SAP export and "
            "upload that instead.",
            _NOTE_FONT,
        ),
    ]


def _width_for(value) -> int:
    """How wide this cell wants to be, in Excel's character units."""
    return len(str(value)) if value is not None else 0


def _measure(head: list) -> tuple[dict[int, int], int]:
    """
    How wide each column wants to be, from the rows already in hand.

    The header is given room for the filter arrow it will carry, which otherwise sits on top
    of the last two characters of every heading.
    """
    widths: dict[int, int] = {}
    columns = 0
    for index, row in enumerate(head):
        for column, value in enumerate(row, start=1):
            wanted = _width_for(value) + (3 if index == 0 else 1)
            widths[column] = max(widths.get(column, 0), wanted)
            columns = max(columns, column)
    return widths, columns


def build_snapshot(upload: DataUploadLog, on_row=None) -> io.BytesIO:
    """
    Rebuild an upload as a protected, readable workbook.

    Only the first sheet is carried. Every master this platform reads is one sheet of rows; a
    second sheet in a SAP export has never been data, and copying one would put something in
    an admin's hands that the importer itself ignored.

    `on_row(written, total)` is called as the copy proceeds, if given. This is the only honest
    measure of the wait: the Member master is a hundred thousand rows and takes most of a
    minute to rebuild, and until it is finished there is not a single byte to send. A client
    told nothing during that draws a bar sitting at zero and then filling in one frame, which
    reads as a broken download rather than a slow one.
    """
    upload.file.seek(0)
    source = load_workbook(upload.file, read_only=True, data_only=True)
    try:
        source_sheet = source[source.sheetnames[0]]
        rows = source_sheet.iter_rows(values_only=True)
        source_rows = source_sheet.max_row or 0

        # Held, measured, then written. The sheet cannot take a width once a row has been
        # appended, so the opening rows have to be in hand before anything is written — and
        # re-opening the file to get them costs more than keeping four hundred of them.
        head = list(islice(rows, WIDTH_SAMPLE_ROWS + 1))
        widths, columns = _measure(head)

        banner = _provenance(upload)
        header_row = len(banner) + 2

        book = Workbook(write_only=True)
        sheet = book.create_sheet(title="Data")

        sheet.protection.sheet = True
        sheet.protection.password = SHEET_PASSWORD
        # Sorting and filtering are how a spreadsheet gets read, and neither changes a value.
        # Locking them would make the file harder to use without making it any safer — the
        # cells are what must not move, and they stay locked.
        sheet.protection.autoFilter = False
        sheet.protection.sort = False
        sheet.protection.enable()

        for column in range(1, columns + 1):
            # Bounded both ways. The floor keeps a two-letter header from collapsing; the
            # ceiling stops one long address column from pushing every other column off the
            # screen, which is the failure mode of naive autofit.
            sheet.column_dimensions[get_column_letter(column)].width = min(
                MAX_COLUMN_WIDTH, max(MIN_COLUMN_WIDTH, widths.get(column, 0))
            )

        if columns:
            # The banner and the header stay put while the rows scroll. A header that
            # disappears on row forty is a file an admin scrolls back up in to remember which
            # column they are reading.
            sheet.freeze_panes = f"A{header_row + 1}"
            last_column = get_column_letter(columns)
            # `source_rows` counts the header with the data, so the filter's last row is the
            # header row plus the data beneath it. Where the file does not declare its size
            # the range covers the header alone, which Excel extends down the contiguous block.
            data_rows = max(0, min(source_rows, MAX_SNAPSHOT_ROWS + 1) - 1)
            sheet.auto_filter.ref = f"A{header_row}:{last_column}{header_row + data_rows}"

        for text, font in banner:
            cell = WriteOnlyCell(sheet, value=text)
            cell.font = font
            sheet.append([cell])
        sheet.append([])

        written = 0
        for index, row in enumerate(chain(head, rows)):
            if written >= MAX_SNAPSHOT_ROWS:
                sheet.append(
                    [
                        f"Truncated at {MAX_SNAPSHOT_ROWS:,} rows. The full master is in the "
                        "platform, not in this file."
                    ]
                )
                break

            if index == 0:
                # The SAP export's own header row, given the portal's header band so it reads
                # as a heading when the file opens rather than as the first record.
                cells = []
                for value in row:
                    cell = WriteOnlyCell(sheet, value=value)
                    cell.font = _HEADER_FONT
                    cell.fill = _HEADER_FILL
                    cell.alignment = Alignment(vertical="center")
                    cells.append(cell)
                sheet.append(cells)
            else:
                sheet.append(list(row))
            written += 1

            # Every few hundred rows rather than every row: the reporting is a network hop to
            # the cache, and a hundred thousand of them would cost more than the work being
            # measured. At this interval the Member master reports about four hundred times,
            # which is a bar that moves smoothly and a cost nobody can see.
            if on_row is not None and written % PROGRESS_EVERY_ROWS == 0:
                on_row(written, source_rows)

        if on_row is not None:
            on_row(written, source_rows)

        buffer = io.BytesIO()
        book.save(buffer)
        buffer.seek(0)
        return buffer
    finally:
        # A read-only workbook holds the file open; the importer learned this the hard way.
        source.close()
