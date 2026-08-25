"""
Handing a master back.

The portal's "Download masters" button answers one question: *what is this platform actually
running on*. Everything under test here protects that answer.

  - it must be the upload that **landed**, not merely the most recent row — a queued or failed
    file never became the master, and offering one as "what we are running on" is the specific
    lie the feature exists to prevent;
  - it must arrive **locked**, because a master that reaches somebody subtly altered is worse
    than no file at all;
  - it must say **where it came from**, because a copy with no date on it is the thing that
    gets mistaken for the current one three weeks later;
  - it must arrive **readable**, because a file that opens as a wall of `#####` under a
    header that scrolls away is one an admin re-formats by hand before every use;
  - and it must be built **once**, because rebuilding the Member master is 165 seconds and a
    request held open that long is a proxy timeout rather than a download.

It also has to say how big it is, and say it in a way a cross-origin portal is allowed to
read — that header is the whole difference between a percentage and a stripe.
"""

from __future__ import annotations

import io
import uuid
from unittest import mock

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.accounts.models import Role, User
from apps.masterdata import snapshots, tasks
from apps.masterdata.models import DataUploadLog
from apps.masterdata.snapshots import (
    MAX_COLUMN_WIDTH,
    MIN_COLUMN_WIDTH,
    SNAPSHOT_VERSION,
)

pytestmark = pytest.mark.django_db


@pytest.fixture
def admin(db):
    return User.objects.create_user(
        username="admin-snapshots",
        password="a-long-enough-password",
        full_name="Snapshot Admin",
        role=Role.ADMIN,
    )


@pytest.fixture
def admin_client(admin):
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {RefreshToken.for_user(admin).access_token}")
    return client


def a_workbook(rows) -> bytes:
    book = Workbook()
    sheet = book.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


@pytest.fixture
def upload(db, admin):
    """One stored upload, in whatever state the case needs."""

    def _make(
        upload_type=DataUploadLog.UploadType.MPP,
        status=DataUploadLog.Status.COMPLETED,
        file_name="All MPP.xlsx",
        rows=(("MPP Code", "MPPName"), ("001302", "ALI PUR"), ("001303", "BAROLI")),
    ):
        return DataUploadLog.objects.create(
            upload_type=upload_type,
            file_name=file_name,
            file=SimpleUploadedFile(file_name, a_workbook(rows)),
            uploaded_by=admin,
            status=status,
            total_rows=len(rows) - 1,
            success_rows=len(rows) - 1,
            finished_at=timezone.now(),
        )

    return _make


def sheet_of(response):
    book = load_workbook(io.BytesIO(response.content))
    return book[book.sheetnames[0]]


def test_the_download_is_the_last_upload_that_landed(admin_client, upload):
    upload(file_name="Old MPP.xlsx")
    upload(file_name="Current MPP.xlsx")
    # Queued after both, and it is not the master: nothing has read it yet.
    upload(file_name="Still Going.xlsx", status=DataUploadLog.Status.QUEUED)
    # Failed after that, and it is not the master either — it never became one.
    upload(file_name="Rejected.xlsx", status=DataUploadLog.Status.FAILED)

    response = admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")

    assert response.status_code == 200
    banner = sheet_of(response).cell(2, 1).value
    assert "Current MPP.xlsx" in banner
    assert "Still Going" not in banner
    assert "Rejected" not in banner


def test_a_partial_import_still_counts_as_landed(admin_client, upload):
    # Its good rows are in force and its bad ones already have a report of their own.
    upload(file_name="Half Good.xlsx", status=DataUploadLog.Status.COMPLETED_WITH_ERRORS)

    response = admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")

    assert "Half Good.xlsx" in sheet_of(response).cell(2, 1).value


def test_the_workbook_is_locked(admin_client, upload):
    # Not a seal — the protection can be removed by anyone who means to. It stops the
    # accident, which is the whole risk: a master altered on its way to somebody.
    upload()

    sheet = sheet_of(admin_client.get("/api/v1/admin/uploads/snapshots/mpp/"))

    assert sheet.protection.sheet is True
    assert sheet.protection.password is not None


def test_the_file_says_where_it_came_from(admin_client, upload):
    upload(file_name="All MPP.xlsx")

    sheet = sheet_of(admin_client.get("/api/v1/admin/uploads/snapshots/mpp/"))

    assert "MPP / Sahayak Master" in sheet.cell(1, 1).value
    assert "All MPP.xlsx" in sheet.cell(2, 1).value
    assert "Snapshot Admin" in sheet.cell(2, 1).value
    # And that editing it achieves nothing, before anybody tries.
    assert "upload that instead" in sheet.cell(3, 1).value


def test_every_row_of_the_master_survives_the_rebuild(admin_client, upload):
    upload(rows=(("MPP Code", "MPPName"), ("001302", "ALI PUR"), ("001303", "BAROLI")))

    sheet = sheet_of(admin_client.get("/api/v1/admin/uploads/snapshots/mpp/"))

    # Three banner lines and a spacer, then the SAP header, then the data.
    assert sheet.cell(5, 1).value == "MPP Code"
    assert [sheet.cell(row, 1).value for row in (6, 7)] == ["001302", "001303"]


def test_a_master_nobody_has_uploaded_is_a_404(admin_client, upload):
    upload(upload_type=DataUploadLog.UploadType.MPP)

    assert admin_client.get("/api/v1/admin/uploads/snapshots/member/").status_code == 404


def test_the_assignment_workbook_is_not_a_master(admin_client, upload):
    # It is generated by this portal rather than uploaded from SAP, and it already has its own
    # round-trip download. Offering it here would be a second door to a different thing.
    upload(upload_type=DataUploadLog.UploadType.ASSIGNMENT)

    assert admin_client.get("/api/v1/admin/uploads/snapshots/assignment/").status_code == 404


def test_the_listing_says_which_masters_have_a_copy(admin_client, upload):
    upload(upload_type=DataUploadLog.UploadType.MPP, file_name="All MPP.xlsx")
    upload(upload_type=DataUploadLog.UploadType.MEMBER, status=DataUploadLog.Status.FAILED)

    body = admin_client.get("/api/v1/admin/uploads/snapshots/").json()
    by_type = {row["upload_type"]: row for row in body["results"]}

    assert by_type["mpp"]["available"] is True
    assert by_type["mpp"]["file_name"] == "All MPP.xlsx"
    # A failed upload leaves the master with nothing to hand back, and the screen has to be
    # able to say so rather than offering a button that 404s.
    assert by_type["member"]["available"] is False
    assert by_type["mait"]["available"] is False


def test_a_mait_cannot_download_a_master(db, mait, upload):
    upload()
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {RefreshToken.for_user(mait.user).access_token}")

    assert client.get("/api/v1/admin/uploads/snapshots/mpp/").status_code == 403
    assert client.get("/api/v1/admin/uploads/snapshots/").status_code == 403


# --- readable on arrival -------------------------------------------------------------------


def test_the_columns_are_sized_to_what_is_in_them(admin_client, upload):
    # Naive autofit is not the goal — one long address column that pushes every other column
    # off the screen is as unusable as a column of hashes. Measured, then bounded.
    upload(
        rows=(
            ("MPP Code", "MPPName", "Address"),
            ("001302", "ALI PUR", "A" * 200),
            ("001303", "BAROLI", "short"),
        )
    )

    sheet = sheet_of(admin_client.get("/api/v1/admin/uploads/snapshots/mpp/"))

    code = sheet.column_dimensions["A"].width
    address = sheet.column_dimensions["C"].width

    assert code >= MIN_COLUMN_WIDTH
    # "MPP Code" plus room for the filter arrow, not the width of the word "Address".
    assert code >= len("MPP Code")
    assert address == MAX_COLUMN_WIDTH


def test_a_narrow_column_still_gets_a_usable_width(admin_client, upload):
    upload(rows=(("ID",), ("1",), ("2",)))

    sheet = sheet_of(admin_client.get("/api/v1/admin/uploads/snapshots/mpp/"))

    assert sheet.column_dimensions["A"].width == MIN_COLUMN_WIDTH


def test_the_header_stays_put_and_can_be_filtered(admin_client, upload):
    upload()

    sheet = sheet_of(admin_client.get("/api/v1/admin/uploads/snapshots/mpp/"))

    # Three banner lines and a spacer put the header on row 5, so the freeze is below it.
    assert sheet.freeze_panes == "A6"
    assert sheet.auto_filter.ref == "A5:B7"

    # Sorting and filtering are reading, not editing, and stay usable on a locked sheet. The
    # cells themselves are still protected.
    assert sheet.protection.sheet is True
    assert sheet.protection.autoFilter is False
    assert sheet.protection.sort is False


def test_the_length_is_sent_and_is_readable_cross_origin(admin_client, upload):
    """
    What a real progress bar needs from the server.

    A length, so the client draws a fraction rather than a stripe that only says "working" —
    and permission to read it, because on the development path the portal is on :8080 and the
    API on :8000, and `Content-Length` is not a CORS-safelisted response header. Without the
    expose header the browser hides it and the percentage quietly disappears.

    Sending it in chunks was tried and reverted: `runserver` is wsgiref and does a blocking
    write per chunk, which turned a 482 KB file into a 12.6-second download.
    """
    upload(
        rows=[("MPP Code", "MPPName")]
        + [(f"{n:06d}", uuid.uuid4().hex) for n in range(2000)]
    )

    response = admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")

    assert int(response["Content-Length"]) == len(response.content)
    exposed = response["Access-Control-Expose-Headers"]
    assert "Content-Length" in exposed
    assert "Content-Disposition" in exposed


# --- built once ----------------------------------------------------------------------------
#
# Rebuilding the Member master measures at 165 seconds for a 27 MB file. On demand that is a
# request held open for most of three minutes; behind nginx it is a proxy timeout. So the
# workbook is kept against the upload it was built from — a row that never changes, because a
# corrected master arrives as a new upload.


def test_the_workbook_is_kept_and_not_rebuilt(admin_client, upload):
    log = upload()

    first = admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")
    log.refresh_from_db()
    assert log.snapshot_file, "the copy is kept against the upload it came from"
    assert log.snapshot_version == SNAPSHOT_VERSION

    # The second download must not touch the builder at all.
    with mock.patch("apps.masterdata.snapshots.build_snapshot") as build:
        second = admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")

    build.assert_not_called()
    assert second.content == first.content


def test_a_formatting_change_reaches_the_copies_people_download(admin_client, upload):
    # The one way a kept file can go stale: the rules that made it changed. A version behind
    # the code is rebuilt rather than served, or a formatting fix lands for new uploads and
    # silently never reaches the masters anybody actually opens.
    log = upload()
    admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")
    log.refresh_from_db()

    DataUploadLog.objects.filter(pk=log.pk).update(snapshot_version=SNAPSHOT_VERSION - 1)

    with mock.patch(
        "apps.masterdata.snapshots.build_snapshot", wraps=snapshots.build_snapshot
    ) as build:
        admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")

    build.assert_called_once()
    log.refresh_from_db()
    assert log.snapshot_version == SNAPSHOT_VERSION


def test_a_copy_missing_from_storage_is_rebuilt_rather_than_a_500(admin_client, upload):
    # The row says there is a file and storage disagrees: a bucket lifecycle rule, a database
    # restored against a fresh volume. The source is still here and the copy was only ever
    # about speed, so this rebuilds.
    log = upload()
    admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")
    log.refresh_from_db()
    log.snapshot_file.storage.delete(log.snapshot_file.name)

    response = admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")

    assert response.status_code == 200
    assert sheet_of(response).cell(5, 1).value == "MPP Code"


def test_the_import_builds_the_copy_before_anybody_asks(db, admin, upload):
    # So the first download is as quick as the tenth. The import already ran for minutes; this
    # is a fraction of it, and it is the difference between a download and a timeout.
    log = upload(status=DataUploadLog.Status.QUEUED)
    assert not log.snapshot_file

    tasks._warm_snapshot(log)

    log.refresh_from_db()
    assert log.snapshot_file
    assert log.snapshot_version == SNAPSHOT_VERSION


def test_a_failed_warm_up_does_not_fail_the_import(db, upload):
    """
    The import is the job; the copy is a convenience.

    An admin who cannot download a spreadsheet is in a better position than one whose
    105,000-row import is reported as failed because a spreadsheet could not be written — and
    the download path rebuilds on demand anyway, so this costs time and nothing else.
    """
    log = upload()

    with mock.patch(
        "apps.masterdata.tasks.store_snapshot", side_effect=OSError("disk full")
    ):
        tasks._warm_snapshot(log)  # must not raise

    log.refresh_from_db()
    assert not log.snapshot_file


def test_the_listing_says_which_masters_are_ready(admin_client, upload):
    # A button that looks identical whether it will answer in a second or in three minutes is
    # a button an operator gives up on.
    upload(upload_type=DataUploadLog.UploadType.MPP)

    before = admin_client.get("/api/v1/admin/uploads/snapshots/").json()["results"]
    assert {row["upload_type"]: row["ready"] for row in before}["mpp"] is False

    admin_client.get("/api/v1/admin/uploads/snapshots/mpp/")

    after = admin_client.get("/api/v1/admin/uploads/snapshots/").json()["results"]
    assert {row["upload_type"]: row["ready"] for row in after}["mpp"] is True
