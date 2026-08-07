"""
Mait ↔ MPP assignment sheet (SRS §6.2.2).

An MPP is the village collection point a Mait covers — the area marker the whole app scopes
on. Getting a row wrong moves a Mait's members, their animals and their permission to record
an AI event at all, so what these tests guard is what the importer refuses.

The pipeline itself (progress, partial success, error rows) is the shared one every SAP
upload uses; this covers the handler's own rules.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.accounts.models import Role, User
from apps.masterdata.models import MPP, DataUploadLog, Mait
from apps.masterdata.tasks import _upsert_assignment
from apps.masterdata.templates_xlsx import build_assignment_workbook

pytestmark = pytest.mark.django_db

HEADERS = ["mpp code", "mait vendor", "mait name", "mobile no"]


def auth(user):
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {RefreshToken.for_user(user).access_token}")
    return client


@pytest.fixture
def admin_client(db):
    return auth(
        User.objects.create_user(
            username="admin-assign",
            password="a-long-enough-password",
            full_name="Admin",
            role=Role.ADMIN,
        )
    )


@pytest.fixture
def mait_client(mait):
    user = User.objects.create_user(username="mait-assign", full_name=mait.name, role=Role.MAIT)
    mait.user = user
    mait.save(update_fields=["user"])
    return auth(user)


class Context:
    """Stands in for ImportContext — only the duplicate check is used by this handler."""

    def __init__(self):
        self.seen = set()

    def check_duplicate(self, key):
        if key in self.seen:
            raise ValueError(f"Duplicate key '{key}' appears more than once in this file.")
        self.seen.add(key)


def row(mpp_code="", vendor="", name="", mobile=""):
    return dict(zip(HEADERS, [mpp_code, vendor, name, mobile], strict=True))


def apply(**kwargs):
    _upsert_assignment(row(**kwargs), Context())


class TestAssigning:
    def test_an_mpp_moves_to_the_named_mait(self, mpp, mait):
        other = Mait.objects.create(sahayak_vendor_code="5500000099", name="NEW MAIT")

        apply(mpp_code=mpp.mpp_code, vendor=other.sahayak_vendor_code)

        mpp.refresh_from_db()
        assert mpp.mait_id == other.id

    def test_a_blank_sahayak_unassigns(self, mpp, mait):
        assert mpp.mait_id == mait.id

        apply(mpp_code=mpp.mpp_code)

        # Not a missing value — it is how an MPP is taken off a Mait who has left.
        mpp.refresh_from_db()
        assert mpp.mait_id is None

    def test_an_unknown_mait_is_created_when_the_row_names_them(self, mpp):
        apply(mpp_code=mpp.mpp_code, vendor="5500000123", name="NEW SAHAYAK", mobile="9876543210")

        created = Mait.objects.get(sahayak_vendor_code="5500000123")
        assert created.name == "NEW SAHAYAK"
        assert created.mobile_no == "9876543210"

    def test_a_retired_record_cannot_be_given_an_mpp(self, mpp, mait):
        mait.is_active = False
        mait.save(update_fields=["is_active"])

        # Almost always a Sahayak code pasted in from the MPP master. Those staff a collection
        # point rather than covering it, and were retired for that reason.
        with pytest.raises(ValueError, match="retired record"):
            apply(mpp_code=mpp.mpp_code, vendor=mait.sahayak_vendor_code)

    def test_an_unknown_mait_with_no_name_is_refused(self, mpp):
        with pytest.raises(ValueError, match="no name"):
            apply(mpp_code=mpp.mpp_code, vendor="5500000404")

        # A nameless row nobody can identify afterwards is worse than a rejected one.
        assert not Mait.objects.filter(sahayak_vendor_code="5500000404").exists()


class TestRefusals:
    def test_an_unknown_mpp_is_refused_rather_than_created(self, mait):
        before = MPP.objects.count()

        with pytest.raises(ValueError, match="No MPP with code"):
            apply(mpp_code="NOPE-123", vendor=mait.sahayak_vendor_code)

        # MPPs come from SAP. A typo must read as a bad row, not bring a new village into
        # existence for a Mait to stare at.
        assert MPP.objects.count() == before

    def test_a_blank_mpp_code_is_refused(self, mait):
        with pytest.raises(ValueError, match="blank"):
            apply(vendor=mait.sahayak_vendor_code)

    def test_the_same_mpp_twice_in_one_file_is_refused(self, mpp, mait):
        context = Context()
        _upsert_assignment(row(mpp.mpp_code, mait.sahayak_vendor_code), context)

        # One MPP has one Mait, so a repeat is a contradiction: the later row would silently
        # win and the count would still look clean.
        with pytest.raises(ValueError, match="more than once"):
            _upsert_assignment(row(mpp.mpp_code, mait.sahayak_vendor_code), context)

    def test_an_unparseable_mobile_is_refused(self, mpp, mait):
        with pytest.raises(ValueError, match="not a usable Indian mobile number"):
            apply(mpp_code=mpp.mpp_code, vendor=mait.sahayak_vendor_code, mobile="12345")

        # The number is the Mait's only way into the app and where their OTP goes, so a value
        # we could not parse is an error rather than a silent blank.


class TestLeavingThingsAlone:
    def test_a_blank_mobile_does_not_wipe_the_one_on_record(self, mpp, mait):
        mait.mobile_no = "9876500001"
        mait.save(update_fields=["mobile_no"])

        apply(mpp_code=mpp.mpp_code, vendor=mait.sahayak_vendor_code, name=mait.name)

        # This is an assignment sheet, not a replacement for the Sahayak master. Clearing the
        # number would lock a working Mait out of the app.
        mait.refresh_from_db()
        assert mait.mobile_no == "9876500001"

    def test_a_supplied_mobile_does_update(self, mpp, mait):
        apply(mpp_code=mpp.mpp_code, vendor=mait.sahayak_vendor_code, mobile="98765 43210")

        mait.refresh_from_db()
        assert mait.mobile_no == "9876543210"

    def test_a_float_looking_code_still_matches(self, mpp, mait):
        # Real MPP codes are numeric ("001302"), and openpyxl hands those back as floats —
        # so the code arrives as "1302.0" and would match nothing without the strip.
        MPP.objects.filter(pk=mpp.pk).update(mpp_code="001302")

        apply(mpp_code="001302.0", vendor=mait.sahayak_vendor_code)

        mpp.refresh_from_db()
        assert mpp.mait_id == mait.id


class TestTemplate:
    def test_it_carries_the_current_mapping(self, mpp, mait):
        workbook = build_assignment_workbook()
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        # "Mait", not "Sahayak" — the MPP master's Sahayak column is a different person.
        assert headers == [
            "MPP Code",
            "MPP Name",
            "District",
            "Mait Vendor",
            "Mait Name",
            "Mobile No",
        ]

        codes = {sheet.cell(row=r, column=1).value: r for r in range(2, sheet.max_row + 1)}
        assert mpp.mpp_code in codes
        assert sheet.cell(row=codes[mpp.mpp_code], column=4).value == mait.sahayak_vendor_code

    def test_unassigned_mpps_are_included(self, mpp):
        MPP.objects.filter(pk=mpp.pk).update(mait=None)

        sheet = build_assignment_workbook().active
        rows = {sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)}

        # They are the rows most likely to need filling; omitting them would hide the work.
        assert mpp.mpp_code in rows

    def test_a_round_trip_of_the_untouched_template_changes_nothing(self, mpp, mait):
        stream = io.BytesIO()
        build_assignment_workbook().save(stream)
        stream.seek(0)
        sheet = load_workbook(stream).active

        context = Context()
        for line in range(2, sheet.max_row + 1):
            _upsert_assignment(
                {
                    "mpp code": sheet.cell(row=line, column=1).value,
                    "mait vendor": sheet.cell(row=line, column=4).value,
                    "mait name": sheet.cell(row=line, column=5).value,
                    "mobile no": sheet.cell(row=line, column=6).value,
                },
                context,
            )

        # The headers the template writes are the ones the importer reads back, so a file that
        # goes out and comes back unedited is a no-op rather than a validation failure.
        mpp.refresh_from_db()
        assert mpp.mait_id == mait.id


class TestEndpoints:
    def test_the_template_downloads_as_a_workbook(self, admin_client):
        response = admin_client.get("/api/v1/admin/uploads/assignment-template/")

        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]
        assert "attachment" in response["Content-Disposition"]

    def test_a_mait_cannot_download_it(self, mait_client):
        assert mait_client.get("/api/v1/admin/uploads/assignment-template/").status_code == 403

    def test_an_upload_is_queued_rather_than_parsed_inline(self, admin_client, settings):
        settings.CELERY_TASK_ALWAYS_EAGER = True
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["MPP Code", "Sahayak Vendor"])
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        stream.name = "assignments.xlsx"

        response = admin_client.post(
            "/api/v1/admin/uploads/assignments/", {"file": stream}, format="multipart"
        )

        assert response.status_code == 202, response.json()
        assert response.json()["upload_type"] == DataUploadLog.UploadType.ASSIGNMENT


class TestInlineEdit:
    """
    Correcting one row without a spreadsheet.

    Most days the change is one Sahayak's number or one village moving between two Maits, and
    downloading three thousand rows to alter one of them is how mistakes get made.
    """

    BASE = "/api/v1/admin/users/maits"

    def test_the_mobile_can_be_corrected(self, admin_client, mait):
        response = admin_client.patch(
            f"{self.BASE}/{mait.sahayak_vendor_code}/", {"mobile_no": "98765 43210"}, format="json"
        )

        assert response.status_code == 200, response.json()
        mait.refresh_from_db()
        assert mait.mobile_no == "9876543210"

    def test_a_bad_mobile_is_refused(self, admin_client, mait):
        response = admin_client.patch(
            f"{self.BASE}/{mait.sahayak_vendor_code}/", {"mobile_no": "12345"}, format="json"
        )

        assert response.status_code == 400

    def test_a_number_another_mait_uses_is_refused(self, admin_client, mait):
        Mait.objects.create(sahayak_vendor_code="5500009999", name="OTHER", mobile_no="9876500002")

        response = admin_client.patch(
            f"{self.BASE}/{mait.sahayak_vendor_code}/", {"mobile_no": "9876500002"}, format="json"
        )

        # Two Maits on one number means an OTP reaching the wrong person, and the second of
        # them can never sign in.
        assert response.status_code == 400

    def test_coverage_is_a_set_not_an_addition(self, admin_client, mait, mpp):
        keep = MPP.objects.create(mpp_code="900001", mpp_name="KEEP", mait=mait)

        response = admin_client.patch(
            f"{self.BASE}/{mait.sahayak_vendor_code}/",
            {"mpp_codes": [keep.mpp_code]},
            format="json",
        )

        assert response.status_code == 200, response.json()
        # The one left out is one they no longer work, and must stop seeing.
        mpp.refresh_from_db()
        assert mpp.mait_id is None
        keep.refresh_from_db()
        assert keep.mait_id == mait.id

    def test_an_unknown_mpp_is_refused(self, admin_client, mait):
        response = admin_client.patch(
            f"{self.BASE}/{mait.sahayak_vendor_code}/", {"mpp_codes": ["NOPE"]}, format="json"
        )

        assert response.status_code == 400
        assert "NOPE" in str(response.json()["errors"])

    def test_omitting_a_field_leaves_it_alone(self, admin_client, mait, mpp):
        mait.mobile_no = "9876500003"
        mait.save(update_fields=["mobile_no"])

        admin_client.patch(
            f"{self.BASE}/{mait.sahayak_vendor_code}/", {"name": "RENAMED"}, format="json"
        )

        # A screen sending the whole object back would wipe whatever it had not loaded.
        mait.refresh_from_db()
        assert mait.mobile_no == "9876500003"
        assert mait.mpps.count() == 1

    def test_a_mait_cannot_edit_the_roster(self, mait_client, mait):
        response = mait_client.patch(
            f"{self.BASE}/{mait.sahayak_vendor_code}/", {"mobile_no": "9876543211"}, format="json"
        )

        assert response.status_code == 403
