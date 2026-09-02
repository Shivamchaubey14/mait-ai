"""
The non-member roster, taken away as a workbook (W10b).

Three things are worth holding still here, and they are the three that would go wrong quietly.

The file has to be **the screen** — same filters, same rows — because an operator who narrows
to the farmers with no card and then exports is entitled to that list and not the whole
population. It has to carry **full Aadhaar and mobile numbers**, because verifying a farmer
against her card is the one job it exists for and a masked number cannot be checked against
anything — so the numbers are pinned here, and so is the audit entry that records who took
them. And the digits have to survive Excel: an Aadhaar that arrives as `1.23457E+11`, or an
MPP code as `1302` instead of `001302`, is a number somebody would verify against and get
wrong.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.accounts.models import Role, User
from apps.core.models import AuditLog
from apps.masterdata.models import NonMember

pytestmark = pytest.mark.django_db

BASE = "/api/v1"
EXPORT = f"{BASE}/admin/non-members/export/"


def auth(api_client, user):
    api_client.credentials(HTTP_AUTHORIZATION=f"Bearer {RefreshToken.for_user(user).access_token}")
    return api_client


def sheet_of(response):
    workbook = load_workbook(io.BytesIO(response.content))
    return workbook["Non-members"]


def headers(sheet):
    """Row 2 — row 1 is the provenance banner."""
    return [cell.value for cell in sheet[2]]


def rows_of(sheet):
    """Every data row as a dict keyed by column name."""
    names = headers(sheet)
    out = []
    for line in sheet.iter_rows(min_row=3, values_only=True):
        if line[0] is None:
            continue
        out.append(dict(zip(names, line, strict=False)))
    return out


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def admin_user(db):
    """Shadows pytest-django's fixture of the same name: this one is an Admin, not a superuser."""
    return User.objects.create_user(
        username="export-admin", password="pw-for-tests-only", full_name="Admin", role=Role.ADMIN
    )


@pytest.fixture
def mait_user(db, mait):
    user = User.objects.create_user(username="export-mait", full_name=mait.name, role=Role.MAIT)
    mait.user = user
    mait.save(update_fields=["user"])
    return user


@pytest.fixture
def registered(api_client, mait_user, mpp):
    def make(name, mobile, aadhaar, consent=True):
        response = auth(api_client, mait_user).post(
            f"{BASE}/non-members/",
            {
                "name": name,
                "father_husband_name": "Ram Singh",
                "relation": "husband",
                "mobile_no": mobile,
                "mpp": mpp.id,
                "aadhar_no": aadhaar,
                "consent": consent,
            },
            format="json",
        )
        assert response.status_code == 201, response.json()
        return NonMember.objects.get(pk=response.json()["id"])

    return make


class TestTheFile:
    def test_it_comes_back_as_a_workbook(self, api_client, admin_user, registered):
        registered("Radha Singh", "9876543210", "111122223333")

        response = auth(api_client, admin_user).get(EXPORT)

        assert response.status_code == 200
        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Dated, because these get mailed around and compared; two downloads a week apart
        # should not be the same filename to the person holding both.
        assert "non-members-" in response["Content-Disposition"]
        assert response["Content-Disposition"].endswith('.xlsx"')

    def test_one_row_per_farmer(self, api_client, admin_user, registered):
        registered("Radha Singh", "9876543210", "111122223333")
        registered("Sunita Devi", "9876543211", "444455556666")

        rows = rows_of(sheet_of(auth(api_client, admin_user).get(EXPORT)))

        assert len(rows) == 2
        assert {row["Name"] for row in rows} == {"Radha Singh", "Sunita Devi"}

    def test_her_record_travels_with_her(self, api_client, admin_user, registered, mpp, mait_user):
        """The columns that make a row auditable without opening it."""
        registered("Radha Singh", "9876543210", "111122223333")

        row = rows_of(sheet_of(auth(api_client, admin_user).get(EXPORT)))[0]

        assert row["Household"] == "Ram Singh"
        assert row["Relation"] == "Husband"
        assert row["MPP code"] == mpp.mpp_code
        assert row["MPP name"] == mpp.mpp_name
        assert row["Registered by"] == mait_user.mait_profile.name
        # Registered but never served, so both counts are honestly zero rather than blank.
        assert row["Animals on file"] == 0
        assert row["AI events"] == 0

    def test_the_card_column_names_the_gap(self, api_client, admin_user, registered):
        """ "Missing" and "Back only" are different jobs, so they are not one word."""
        registered("Radha Singh", "9876543210", "111122223333")

        row = rows_of(sheet_of(auth(api_client, admin_user).get(EXPORT)))[0]

        assert row["Card"] == "Missing"

    def test_a_farmer_with_no_consent_says_so(self, api_client, admin_user, registered):
        registered("Radha Singh", "9876543210", "111122223333", consent=False)

        row = rows_of(sheet_of(auth(api_client, admin_user).get(EXPORT)))[0]

        assert row["Consent"] == "Not captured"


class TestWhatLeavesThePlatform:
    def test_aadhaar_and_mobile_come_through_in_full(self, api_client, admin_user, registered):
        """
        The exception to the rule the other exports keep, and the reason this file exists: an
        operator checks the number on the row against the number on the card. A masked one
        cannot be checked against anything.
        """
        registered("Radha Singh", "9876543210", "111122223333")

        row = rows_of(sheet_of(auth(api_client, admin_user).get(EXPORT)))[0]

        assert row["Aadhaar"] == "111122223333"
        assert row["Mobile"] == "9876543210"

    def test_the_sheet_warns_what_it_carries(self, api_client, admin_user, registered):
        """
        The person who opens it decides where it gets stored and who it goes to, and cannot
        make that call without being told what is in it.
        """
        registered("Radha Singh", "9876543210", "111122223333")

        banner = sheet_of(auth(api_client, admin_user).get(EXPORT))["A1"].value

        assert "AADHAAR" in banner
        assert "personal data" in banner.lower()

    def test_the_export_is_logged_against_the_admin(self, api_client, admin_user, registered):
        registered("Radha Singh", "9876543210", "111122223333")

        auth(api_client, admin_user).get(f"{EXPORT}?no_card=true")

        entry = AuditLog.objects.filter(entity_id="non_members_export").latest("id")
        assert entry.action == AuditLog.Action.PII_ACCESS
        # The filters too: which rows left matters as much as that some did.
        assert entry.meta_json["filters"]["no_card"] == "true"

    def test_a_mait_cannot_export_the_whole_population(self, api_client, mait_user, registered):
        """A Mait's own list is scoped to what they registered; this is everybody's."""
        registered("Radha Singh", "9876543210", "111122223333")

        assert auth(api_client, mait_user).get(EXPORT).status_code == 403


class TestTheFileIsTheScreen:
    def test_a_search_narrows_the_file(self, api_client, admin_user, registered):
        registered("Radha Singh", "9876543210", "111122223333")
        registered("Sunita Devi", "9876543211", "444455556666")

        rows = rows_of(sheet_of(auth(api_client, admin_user).get(f"{EXPORT}?search=Sunita")))

        assert [row["Name"] for row in rows] == ["Sunita Devi"]

    def test_the_no_card_queue_exports_as_the_queue(self, api_client, admin_user, registered):
        """The filter an operator opens this screen twice for."""
        registered("Radha Singh", "9876543210", "111122223333")
        registered("Sunita Devi", "9876543211", "444455556666")

        rows = rows_of(sheet_of(auth(api_client, admin_user).get(f"{EXPORT}?no_card=true")))

        assert len(rows) == 2  # neither has a card yet

    def test_it_is_the_whole_set_not_the_page(self, api_client, admin_user, registered):
        """
        `limit` and `offset` are how the table pages, not part of the query. Honoured here,
        the file would be the twenty-five rows that happened to be on screen — complete-looking
        and wrong.
        """
        for index in range(3):
            registered(f"Farmer {index}", f"987654321{index}", f"11112222333{index}")

        rows = rows_of(sheet_of(auth(api_client, admin_user).get(f"{EXPORT}?limit=1&offset=0")))

        assert len(rows) == 3


class TestExcelDoesNotEatIt:
    def test_codes_and_numbers_are_text(self, api_client, admin_user, registered):
        """
        An MPP code left as a number loses the leading zero off `001302` and names an MPP that
        does not exist; a mobile becomes `9.87654E+09`. Both are the file coming back wrong
        after a round trip through somebody's laptop.
        """
        registered("Radha Singh", "9876543210", "111122223333")

        sheet = sheet_of(auth(api_client, admin_user).get(EXPORT))
        names = headers(sheet)
        for column in ("Mobile", "Aadhaar", "MPP code", "Mait code", "Registered on"):
            cell = sheet.cell(row=3, column=names.index(column) + 1)
            assert cell.number_format == "@", f"{column} is not text"

    def test_counts_stay_numeric(self, api_client, admin_user, registered):
        """A pivot table has to be able to sum the herd; text would not."""
        registered("Radha Singh", "9876543210", "111122223333")

        sheet = sheet_of(auth(api_client, admin_user).get(EXPORT))
        names = headers(sheet)
        for column in ("Cows", "Buffaloes", "Herd", "Litres/day", "Animals on file"):
            cell = sheet.cell(row=3, column=names.index(column) + 1)
            assert isinstance(cell.value, int | float), f"{column} is not a number"

    def test_the_header_survives_a_scroll(self, api_client, admin_user, registered):
        registered("Radha Singh", "9876543210", "111122223333")

        sheet = sheet_of(auth(api_client, admin_user).get(EXPORT))

        assert sheet.freeze_panes == "A3"
        assert sheet.auto_filter.ref is not None


class TestWhereSheIs:
    """
    The position columns, and the three ways they can lie.

    A `NonMember` holds no coordinates of her own — she is a name, a number and a card, typed
    in a yard — so the only record of where she is, is the pin on an AI event captured at her
    animal. That makes the column useful and also makes it something to be careful with: it
    can be stale, it can come from a photograph taken anywhere, and it can be absent.
    """

    def a_fix(self, non_member, mait, mpp, lat, lng, *, when, source="device"):
        """One completed insemination at her animal, with a GPS pin on it."""
        import uuid

        from apps.ai_events.models import AIEvent
        from apps.animals.models import Animal, AnimalType
        from apps.inventory.models import SemenBatch

        animal = Animal.objects.create(
            owner_type=Animal.OwnerType.NON_MEMBER,
            non_member=non_member,
            animal_type=AnimalType.COW,
            breed="GIR",
        )
        straw = SemenBatch.objects.create(
            unique_straw_no=uuid.uuid4().hex[:20], breed="GIR", is_consumed=True
        )
        return AIEvent.objects.create(
            client_uuid=uuid.uuid4(),
            mait=mait,
            mpp=mpp,
            owner_type=AIEvent.OwnerType.NON_MEMBER,
            non_member=non_member,
            animal=animal,
            semen_batch=straw,
            straw_unique_no=straw.unique_straw_no,
            status=AIEvent.Status.COMPLETED,
            gps_lat=lat,
            gps_lng=lng,
            gps_source=source,
            performed_at=when,
            completed_at=when,
        )

    def test_the_fix_from_her_own_insemination_reaches_the_file(
        self, api_client, admin_user, registered, mait, mpp
    ):
        from django.utils import timezone

        her = registered("Radha Singh", "9876543210", "111122223333")
        self.a_fix(her, mait, mpp, "25.9113176", "82.4823914", when=timezone.now())

        row = rows_of(sheet_of(auth(api_client, admin_user).get(EXPORT)))[0]

        assert row["Latitude"] == pytest.approx(25.9113176)
        assert row["Longitude"] == pytest.approx(82.4823914)
        assert row["Position from"] == "Handset"

    def test_the_newest_fix_wins(self, api_client, admin_user, registered, mait, mpp):
        """
        A farmer moves and her herd moves with her, so the latest visit is the one worth
        driving to. Ordered by when the insemination happened rather than by row id — events
        sync from handsets that were offline, so the last row written is not the last visit
        made.
        """
        from datetime import timedelta

        from django.utils import timezone

        her = registered("Radha Singh", "9876543210", "111122223333")
        now = timezone.now()
        # Written second, performed first: the offline-sync case.
        self.a_fix(her, mait, mpp, "26.1513092", "81.8189077", when=now)
        self.a_fix(her, mait, mpp, "11.1111111", "72.2222222", when=now - timedelta(days=90))

        row = rows_of(sheet_of(auth(api_client, admin_user).get(EXPORT)))[0]

        assert row["Latitude"] == pytest.approx(26.1513092)

    def test_a_pin_off_a_photograph_says_so(self, api_client, admin_user, registered, mait, mpp):
        """
        A device fix is where the handset was standing. A photograph's is whatever EXIF the
        chosen image carried — possibly another district, possibly last year. Presenting the
        second as the first is how a farmer gets mapped to a place nobody ever visited.
        """
        from django.utils import timezone

        her = registered("Radha Singh", "9876543210", "111122223333")
        self.a_fix(her, mait, mpp, "25.9", "82.4", when=timezone.now(), source="photo")

        row = rows_of(sheet_of(auth(api_client, admin_user).get(EXPORT)))[0]

        assert row["Position from"] == "Photograph"

    def test_a_farmer_with_no_fix_is_blank_and_not_a_zero(self, api_client, admin_user, registered):
        """
        0, 0 is a point in the Gulf of Guinea. A column of them would put every unlocated
        farmer on the same island, which is a plausible answer where an obviously missing one
        was wanted.
        """
        registered("Radha Singh", "9876543210", "111122223333")

        row = rows_of(sheet_of(auth(api_client, admin_user).get(EXPORT)))[0]

        assert row["Latitude"] in (None, "")
        assert row["Longitude"] in (None, "")
        assert row["Position from"] in (None, "")
        assert row["Position taken"] in (None, "")

    def test_the_coordinate_keeps_its_precision(
        self, api_client, admin_user, registered, mait, mpp
    ):
        """
        Seven decimal places, the precision the column stores. Excel's general format rounds
        to about five, which is a different place by a metre — no distance at all on a map, and
        enough that a coordinate copied out and pasted back no longer matches its own record.
        """
        from django.utils import timezone

        her = registered("Radha Singh", "9876543210", "111122223333")
        self.a_fix(her, mait, mpp, "25.9113176", "82.4823914", when=timezone.now())

        sheet = sheet_of(auth(api_client, admin_user).get(EXPORT))
        column = headers(sheet).index("Latitude") + 1
        assert sheet.cell(row=3, column=column).number_format == "0.0000000"

    def test_the_card_tint_still_finds_the_right_columns(self, api_client, admin_user, registered):
        """
        The rows the file is opened to work through are tinted, and which cells decide that is
        looked up by column name. Inserting the position columns in front of them used to move
        the two this reads, silently tinting whichever rows happened to land there.
        """
        registered("No consent", "9876543212", "777788889999", consent=False)

        sheet = sheet_of(auth(api_client, admin_user).get(EXPORT))
        assert sheet.cell(row=3, column=1).fill.fgColor.rgb.endswith("FDECEC")
