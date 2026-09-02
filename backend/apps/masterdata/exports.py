"""
The non-member roster, as a workbook the back office can take away (W10b).

The screen it comes from is a review queue: rows typed by a Mait in a yard, on the one path
in this product that ends with cash changing hands. Reviewing them on the screen works while
there are fifty; the moment somebody has to reconcile a district against a milk-payment run,
or sit in a field meeting with the list of farmers whose card was never photographed, they
need it in something they can sort and filter. That is what this is for.

**xlsx rather than the CSV the other two exports produce.** Not a preference — the columns
here are full of things Excel corrupts on the way in from a CSV, and this file carries the
worst of them. A twelve-digit Aadhaar becomes ``1.23457E+11``, a mobile number becomes
``9.19876E+11``, and an MPP code loses the leading zero off ``001302``. A number that arrives
rounded into scientific notation is worse than one that never arrived: it still looks like a
number, and it is the number somebody is about to verify a farmer against. Written as a real
workbook, every one of those cells is typed as text and arrives intact. ``templates_xlsx``
already had to learn this about the assignment sheet.

**Built in memory, not streamed.** The other exports stream because a month of AI events is
tens of thousands of rows; this is bounded by the non-member population, which is farmers
registered by hand — a few thousand, growing at the speed of people. An xlsx is a zip and
cannot be valid until its central directory is written, so there is nothing to stream anyway:
the file has to be finished before its first byte is honest. ``MAX_ROWS`` is the guard against
that assumption quietly ceasing to be true.

**Aadhaar and mobile leave in full, and that is a deliberate exception** to the rule the
AI-event and Pregnancy exports keep. Those two are read to answer questions about the work —
how many inseminations, which villages, who is owed a check — and a masked number costs them
nothing. This file is read to *verify the farmer*: an operator sits with it against the card
images and the registration and confirms that the woman a Mait charged in cash is who the row
says she is. A masked number cannot be checked against anything, which would leave this file
unable to do the one job it exists for.

So the exposure is accepted rather than hidden, and three things hold it in place. The
endpoint is Admin-only and behind the Non-members section, the same gate as the detail screen —
which already returns the card images, so a full number was readable to exactly these people
before this file existed. Every export is written to the audit log with the filters it ran,
so who took which rows away is on the record. And the sheet says on its first row that it
carries them, because the person who opens it in a month is the person who has to decide where
it may be stored (SRS §7, §16).
"""

from __future__ import annotations

from django.db.models import OuterRef, Subquery
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.ai_events.models import AIEvent
from apps.core.timeframe import local_day

#: The population is hand-entered, so this is a long way off. It exists so the decision to
#: build in memory fails loudly if that stops being true, rather than by degrees.
MAX_ROWS = 50_000

#: What `gps_source` means to somebody reading a spreadsheet.
#:
#: The distinction is not pedantry and it is not collapsed. A device fix is where the handset
#: was standing when the Mait captured the insemination. A photograph's fix is whatever EXIF
#: the chosen image happened to carry — it can be anywhere and any time, including a picture
#: taken in another district last year. Presenting the second as though it were the first is
#: how a farmer ends up mapped to a place nobody ever visited, and `ai_events.models` keeps the
#: two apart for exactly this reason.
POSITION_SOURCE = {
    AIEvent.GpsSource.DEVICE: "Handset",
    AIEvent.GpsSource.PHOTO: "Photograph",
}

#: Seven decimal places, the precision the column stores. Left to itself Excel shows a general
#: number rounded to something like 25.91132, which is a different place by about a metre — not
#: enough to matter to a map, but enough that a coordinate copied out of this file and pasted
#: back in no longer matches the record it came from.
COORDINATE_FORMAT = "0.0000000"
COORDINATE_COLUMNS = {"Latitude", "Longitude"}

# The portal's ink, so the file looks like it came from the same product as the assignment
# sheet.
HEADER_FILL = PatternFill("solid", fgColor="253D4E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
# Not a quiet grey note: this line is a warning about what the file holds, and it has to
# survive being skimmed past on the way to the data.
WARN_FONT = Font(color="B42318", bold=True, size=9)
# A row missing either of the two things that keep this path honest, tinted the way the screen
# tints them. These are what the file was downloaded to work through.
GAP_FILL = PatternFill("solid", fgColor="FDECEC")

#: ``(header, width, force_text)``. Codes and numbers are text; counts and litres are genuinely
#: numeric and stay that way, so a pivot table can sum them.
COLUMNS = [
    ("Name", 26, False),
    ("Household", 24, False),
    ("Relation", 12, False),
    ("Mobile", 16, True),
    ("Aadhaar", 18, True),
    ("Card", 12, False),
    ("Consent", 14, False),
    ("MPP code", 12, True),
    ("MPP name", 26, False),
    # Where she actually is, next to the collection point she is filed under — which is the
    # question the MPP column gets asked and cannot answer. An MPP is a village-level pool
    # covering a scatter of households; these four say which household.
    ("Latitude", 13, False),
    ("Longitude", 13, False),
    ("Position from", 14, False),
    ("Position taken", 15, True),
    ("Registered by", 24, False),
    ("Mait code", 14, True),
    ("Cows", 8, False),
    ("Buffaloes", 11, False),
    ("Herd", 8, False),
    ("Litres/day", 11, False),
    ("Animals on file", 15, False),
    ("AI events", 11, False),
    ("Registered on", 15, True),
]


def _at(title: str) -> int:
    """Where a column sits in a row, by name. See `_row`'s use of it."""
    return next(i for i, (name, _w, _t) in enumerate(COLUMNS) if name == title)


def with_last_known_position(queryset):
    """
    Attach each farmer's most recent GPS fix, taken from her own inseminations.

    A `NonMember` carries no coordinates of her own — she is a name, a number and a card, typed
    in a yard — so the only place this platform has ever recorded where she is, is the pin on
    an AI event captured at her animal. That is a good answer and worth exporting: an MPP is a
    village-level pool covering a scatter of households, so "which MPP" has never been able to
    say where to drive.

    The **most recent** one, because a farmer moves and a herd moves with her, and the newest
    fix is the one worth navigating to. Ordered by `performed_at` rather than by row id: events
    sync from handsets that were offline, so the last row written is not the last visit made.

    Four correlated subqueries, applied only where they are wanted. The list screen has no
    position column, and paying for them on every page of fifty rows would buy nothing.
    """
    fixes = AIEvent.objects.filter(
        non_member=OuterRef("pk"), gps_lat__isnull=False, gps_lng__isnull=False
    ).order_by("-performed_at", "-id")
    return queryset.annotate(
        last_lat=Subquery(fixes.values("gps_lat")[:1]),
        last_lng=Subquery(fixes.values("gps_lng")[:1]),
        last_gps_source=Subquery(fixes.values("gps_source")[:1]),
        last_gps_at=Subquery(fixes.values("performed_at")[:1]),
    )


def _card(non_member) -> str:
    """
    The card, in the same four words the screen uses.

    "Missing" and "Back only" are different jobs for the back office — one is a registration to
    redo, the other a photograph to chase — so they are not collapsed into a single word here
    either. ``non-members.js`` draws exactly these.
    """
    front = bool(non_member.aadhar_front_url)
    back = bool(non_member.aadhar_back_url)
    if front and back:
        return "On file"
    if front:
        return "Front only"
    if back:
        return "Back only"
    return "Missing"


def _consent(non_member) -> str:
    """When she agreed to be on file, or the fact that there is no record of her agreeing."""
    if non_member.consent_captured_at:
        return local_day(non_member.consent_captured_at).isoformat()
    return "Not captured"


def _coordinate(non_member, field: str):
    """One half of a fix as a number, or blank where there is none to give."""
    value = getattr(non_member, field, None)
    return float(value) if value is not None else ""


def _position_day(non_member) -> str:
    """The local day the fix was taken, which is what says whether it is still worth using."""
    when = getattr(non_member, "last_gps_at", None)
    return local_day(when).isoformat() if when else ""


def _row(non_member) -> list:
    mait = non_member.created_by_mait
    mpp = non_member.mpp
    return [
        non_member.name,
        non_member.father_husband_name or "Not recorded",
        non_member.get_relation_display() or "",
        # In full, both of them. This file is what the number is checked *against* — see the
        # note at the top of the module about why that is the exception here.
        non_member.mobile_no or "",
        non_member.aadhar_no or "",
        _card(non_member),
        _consent(non_member),
        mpp.mpp_code if mpp else "",
        mpp.mpp_name if mpp else "",
        # Blank, never zero, where there is no fix. Latitude 0, longitude 0 is a point in the
        # Gulf of Guinea, and a column of them would put every unlocated farmer on the same
        # island — a plausible answer is worse than an obviously missing one on a map.
        #
        # Annotated by `with_last_known_position`, and defaulted for the same reason the row
        # counts below are: the builder is also reachable from a shell and from tests, where
        # the annotation may not be there.
        _coordinate(non_member, "last_lat"),
        _coordinate(non_member, "last_lng"),
        POSITION_SOURCE.get(getattr(non_member, "last_gps_source", None), ""),
        _position_day(non_member),
        mait.name if mait else "",
        mait.sahayak_vendor_code if mait else "",
        non_member.cattle_cows,
        non_member.cattle_buffaloes,
        non_member.cattle_total,
        # A Decimal reaches openpyxl as a number, but float() keeps it one rather than letting
        # it arrive as the string repr of a Decimal.
        float(non_member.daily_yield_litres or 0),
        # Annotated by the viewset. Defaulted because the builder is also reachable from a
        # shell and from tests, where the annotation may not be there.
        getattr(non_member, "animal_count", 0) or 0,
        getattr(non_member, "ai_event_count", 0) or 0,
        local_day(non_member.created_at).isoformat(),
    ]


def build_non_member_workbook(queryset) -> Workbook:
    """
    One sheet, one row per farmer, in whatever order the queryset arrived in.

    The order is the caller's on purpose: this is taken from a screen and it should match the
    screen it was taken from, down to the sort. Re-sorting here is how a file and its preview
    quietly stop agreeing.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Non-members"

    # Said on the sheet, not only in this docstring. Somebody opens the file a month later with
    # no idea where it came from, and the person holding it is the person who decides where it
    # gets stored and who it gets sent to — they cannot make that call without being told what
    # is in it.
    stamp = timezone.localdate().isoformat()
    banner = sheet.cell(
        row=1,
        column=1,
        value=(
            f"Non-members registered by Maits in the field · exported {stamp} · "
            "CONTAINS FULL AADHAAR AND MOBILE NUMBERS — handle as personal data"
        ),
    )
    banner.font = WARN_FONT

    for index, (title, width, _force_text) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=2, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width

    line = 3
    for non_member in queryset.iterator(chunk_size=1000):
        values = _row(non_member)
        # Tinted rather than filtered out: the file is the whole queue, and these are the part
        # of it somebody has to act on. Looked up by column name rather than by position —
        # inserting a column used to move the two this reads, silently tinting the wrong rows.
        incomplete = values[_at("Card")] != "On file" or values[_at("Consent")] == "Not captured"
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=line, column=index, value=value)
            title, _width, force_text = COLUMNS[index - 1]
            if force_text:
                cell.number_format = "@"
            elif title in COORDINATE_COLUMNS and value != "":
                cell.number_format = COORDINATE_FORMAT
            if incomplete:
                cell.fill = GAP_FILL
        line += 1
        if line > MAX_ROWS + 2:
            break

    # Below the banner, so a scroll through the roster keeps the column names in view. The
    # filter arrows sit on the header row, because sorting and filtering a list is reading it.
    sheet.freeze_panes = "A3"
    if line > 3:
        sheet.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{line - 1}"
    return workbook


def non_member_workbook_response(queryset) -> HttpResponse:
    workbook = build_non_member_workbook(queryset)
    stamp = timezone.localdate().isoformat()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="non-members-{stamp}.xlsx"'
    workbook.save(response)
    workbook.close()
    return response
