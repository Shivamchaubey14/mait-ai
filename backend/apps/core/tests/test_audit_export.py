"""
Handing the audit trail to an auditor (W19).

The screen is read by somebody who works here. This file is read by somebody who does not,
months later, with no access to the running system — so what these tests pin is not the query
(``test_audit_api`` already owns that) but the three things that make a spreadsheet evidence
rather than a spreadsheet.

**It says where it came from.** Who took it, when, with which filters, and the request id of
the entry the export wrote to the trail on its way out.

**It admits what it is missing.** A file capped at ``MAX_ROWS`` that does not say so is worse
than no file: it looks complete and it is not.

**Taking it is itself recorded.** A record of who reads the record is the whole point of the
record, and an export path that quietly does not write one is the single most valuable gap an
insider could find in this platform.
"""

from __future__ import annotations

import io
from datetime import timedelta

import pytest
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.accounts.models import PortalSection, Role, User
from apps.core import audit_export
from apps.core.models import AuditLog

pytestmark = pytest.mark.django_db

URL = "/api/v1/admin/audit/export/"


@pytest.fixture
def auditor(db):
    return User.objects.create_user(
        username="auditor",
        password="pw-for-tests-only",
        full_name="Asha Verma",
        role=Role.ADMIN,
        portal_sections=[PortalSection.USERS, PortalSection.LOGS],
    )


@pytest.fixture
def client_for(db):
    def _make(user):
        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f"Bearer {RefreshToken.for_user(user).access_token}")
        return client

    return _make


@pytest.fixture
def audit_client(auditor, client_for):
    return client_for(auditor)


def an_entry(**fields):
    """One trail row, placed in the past where asked."""
    minutes = fields.pop("minutes_ago", 5)
    entry = AuditLog.objects.create(
        action=fields.pop("action", AuditLog.Action.CREATE),
        entity_type=fields.pop("entity_type", "ai_event"),
        entity_id=str(fields.pop("entity_id", 1)),
        meta_json=fields.pop("meta", {}),
        **fields,
    )
    AuditLog.objects.filter(pk=entry.pk).update(
        created_at=timezone.now() - timedelta(minutes=minutes)
    )
    entry.refresh_from_db()
    return entry


def workbook(response):
    return load_workbook(io.BytesIO(response.content))


def cells(sheet) -> list[list]:
    """Every row as a plain list, so a test can say what it means without openpyxl noise."""
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def column(sheet, title: str) -> list:
    """One column of the data, by header name rather than by position."""
    rows = cells(sheet)
    index = rows[0].index(title)
    return [row[index] for row in rows[1:]]


def cover_text(book) -> str:
    return "\n".join(
        " ".join(str(value) for value in row if value is not None)
        for row in cells(book["About this export"])
    )


# --------------------------------------------------------------------------------------
# The shape of the file
# --------------------------------------------------------------------------------------
def test_the_workbook_has_the_four_sheets_and_opens_on_the_cover(audit_client):
    an_entry()

    book = workbook(audit_client.get(URL))

    assert book.sheetnames == ["About this export", "Audit trail", "Personal data", "Who did what"]
    # Whatever the reader's Excel remembers about the last sheet they were on, the file has
    # to open on the page that says what it is.
    assert book.active.title == "About this export"


def test_a_row_reads_as_the_sentence_the_screen_shows(audit_client):
    """
    The file and the screen say the same thing about the same row.

    An auditor quoting "Completed AI event 64" at an operator who is looking at different
    words for that row is exactly the confusion this file exists to prevent, so the workbook
    is built through the same `serialise` the screen renders from.
    """
    an_entry(
        action=AuditLog.Action.STATE_CHANGE,
        entity_type="ai_event",
        entity_id=64,
        meta={"to": "completed", "straw": "T0001-HF-0002"},
    )

    sheet = workbook(audit_client.get(URL))["Audit trail"]

    assert column(sheet, "What happened") == ["Completed AI event 64"]
    assert column(sheet, "Record type") == ["AI event"]
    assert column(sheet, "Record") == ["64"]
    # The metadata the sentence did not use, labelled rather than dumped as JSON.
    assert column(sheet, "Details") == ["Straw: T0001-HF-0002"]


def test_a_before_and_after_arrives_as_one_readable_cell(audit_client):
    an_entry(
        action=AuditLog.Action.UPDATE,
        entity_type="mait_payout_scheme",
        entity_id=1,
        meta={"before": {"straw_rate": 40}, "after": {"straw_rate": 45}},
    )

    sheet = workbook(audit_client.get(URL))["Audit trail"]

    assert column(sheet, "Changed") == ["Straw rate: 40 → 45"]


def test_identifiers_are_stored_as_text(audit_client):
    """
    An MPP code is ``001302``, not ``1302``, and a request id is not a number in scientific
    notation. Excel decides that on the way in unless the cell says otherwise.
    """
    an_entry(entity_type="mpp", entity_id="001302")

    sheet = workbook(audit_client.get(URL))["Audit trail"]
    header = cells(sheet)[0]

    for title in ("Date", "Time", "Record", "Username", "IP address", "Request id"):
        assert sheet.cell(row=2, column=header.index(title) + 1).number_format == "@"
    assert column(sheet, "Record") == ["001302"]


# --------------------------------------------------------------------------------------
# Personal data, on its own
# --------------------------------------------------------------------------------------
def test_personal_data_reads_get_their_own_sheet(audit_client):
    """
    SRS §16 asks for a record of who read personal data. "It is in the other sheet, filter for
    it" is not an answer to that question, so the rows are repeated deliberately.
    """
    an_entry(action=AuditLog.Action.CREATE, entity_type="animal", entity_id=7)
    an_entry(
        action=AuditLog.Action.PII_ACCESS,
        entity_type="member",
        entity_id=1837,
        meta={"aadhaar_card_viewed": True},
    )

    book = workbook(audit_client.get(URL))

    assert column(book["Personal data"], "What happened") == [
        "Opened the Aadhaar card on member 1837"
    ]
    # And still present in the full trail, which is the sheet somebody reads end to end.
    assert "Opened the Aadhaar card on member 1837" in column(book["Audit trail"], "What happened")


def test_an_empty_personal_data_sheet_says_so(audit_client):
    """A blank sheet reads as one that failed to load. This one is a finding in its own right."""
    an_entry(action=AuditLog.Action.CREATE, entity_type="animal", entity_id=7)

    sheet = workbook(audit_client.get(URL))["Personal data"]

    assert sheet.cell(row=2, column=1).value == (
        "No personal data was read in the entries this file covers."
    )


def test_the_people_sheet_counts_the_file_by_person(audit_client, auditor, client_for):
    other = User.objects.create_user(
        username="ops",
        password="pw-for-tests-only",
        full_name="Ops Person",
        role=Role.ADMIN,
        portal_sections=[PortalSection.LOGS],
    )
    an_entry(actor=auditor, action=AuditLog.Action.CREATE, entity_id=1, minutes_ago=30)
    an_entry(actor=auditor, action=AuditLog.Action.CREATE, entity_id=2, minutes_ago=20)
    an_entry(
        actor=other,
        action=AuditLog.Action.PII_ACCESS,
        entity_type="member",
        entity_id=9,
        minutes_ago=10,
    )

    sheet = workbook(audit_client.get(URL))["Who did what"]
    people = {row[1]: row for row in cells(sheet)[1:]}

    assert people["auditor"][3] == 2
    assert people["auditor"][4] == 0
    assert people["ops"][3] == 1
    assert people["ops"][4] == 1
    # Whoever read personal data sorts above whoever did not, however few entries they have.
    # An auditor works down from the top of this sheet.
    assert list(people) == ["ops", "auditor"]


# --------------------------------------------------------------------------------------
# Provenance
# --------------------------------------------------------------------------------------
def test_the_cover_says_who_took_it_and_when(audit_client):
    an_entry()

    text = cover_text(workbook(audit_client.get(URL)))

    assert "Asha Verma" in text
    assert "auditor" in text
    assert "Exported at" in text
    assert "CONTAINS PERSONAL DATA" in text
    # The promise the trail makes, said in the file rather than only in this repo.
    assert "append-only" in text


def test_the_cover_describes_the_filters_in_words(audit_client):
    """
    ``action=pii_access`` is what was sent. "Action: Personal data read" is what somebody
    reading the file in a year can act on — the scope of the evidence is the first thing they
    have to establish.
    """
    an_entry(action=AuditLog.Action.PII_ACCESS, entity_type="member", entity_id=1)

    text = cover_text(
        workbook(
            audit_client.get(
                URL, {"action": AuditLog.Action.PII_ACCESS, "date_from": "2026-08-01"}
            )
        )
    )

    assert "Personal data read" in text
    assert "2026-08-01 onwards" in text


def test_an_unfiltered_export_says_it_is_unfiltered(audit_client):
    """Silence about the filters reads as "none applied", and a guess is not provenance."""
    an_entry()

    assert "None — the whole trail" in cover_text(workbook(audit_client.get(URL)))


def test_the_cover_carries_the_request_id_of_its_own_entry(audit_client):
    """
    The thread back to the running system. Somebody holding this file can find that request in
    the live trail and confirm the file was produced when it claims to have been.
    """
    an_entry()

    response = audit_client.get(URL)
    recorded = AuditLog.objects.filter(
        action=AuditLog.Action.PII_ACCESS, entity_id="audit_trail"
    ).latest("created_at")

    assert recorded.request_id
    assert recorded.request_id in cover_text(workbook(response))


# --------------------------------------------------------------------------------------
# Taking the file is itself an event
# --------------------------------------------------------------------------------------
def test_taking_the_export_is_recorded_as_a_personal_data_read(audit_client, auditor):
    an_entry()

    audit_client.get(URL, {"action": AuditLog.Action.LOGIN, "search": "asha"})

    recorded = AuditLog.objects.filter(entity_id="audit_trail").latest("created_at")
    assert recorded.action == AuditLog.Action.PII_ACCESS
    assert recorded.actor == auditor
    # With the filters it ran, so the record says which rows left the building.
    assert recorded.meta_json["filters"] == {"action": AuditLog.Action.LOGIN, "search": "asha"}


def test_the_export_entry_reads_as_a_sentence(audit_client):
    """
    The trail describes its own export in the same words as everything else. `entity_id` here
    is user-facing prose on its way through `describe`, not an internal key.
    """
    audit_client.get(URL)

    recorded = AuditLog.objects.filter(entity_id="audit_trail").latest("created_at")
    from apps.core.audit_api import describe

    assert describe(recorded) == "Exported audit trail"


def test_the_paging_of_the_screen_is_not_the_scope_of_the_file(audit_client):
    """
    `limit` and `offset` are how the table pages, not part of the question. Honouring them
    would hand back the fifty rows that happened to be on screen — a file that looks complete,
    is not, and says nothing about it.
    """
    for index in range(6):
        an_entry(entity_id=index, minutes_ago=index + 1)

    sheet = workbook(audit_client.get(URL, {"limit": 2, "offset": 4}))["Audit trail"]

    assert len(cells(sheet)) - 1 == 6


def test_the_filters_of_the_screen_are_the_scope_of_the_file(audit_client):
    an_entry(action=AuditLog.Action.CREATE, entity_type="animal", entity_id=7)
    an_entry(action=AuditLog.Action.LOGIN, entity_type="user", entity_id=2)

    sheet = workbook(audit_client.get(URL, {"action": AuditLog.Action.LOGIN}))["Audit trail"]

    assert column(sheet, "What happened") == ["Signed in"]


# --------------------------------------------------------------------------------------
# The ceiling
# --------------------------------------------------------------------------------------
def test_a_capped_file_says_on_its_cover_that_it_is_incomplete(audit_client, monkeypatch):
    """
    The trail is the one export here bounded by nothing but time. A file that silently stops
    at the cap is evidence of the wrong thing.
    """
    monkeypatch.setattr(audit_export, "MAX_ROWS", 2)
    for index in range(5):
        an_entry(entity_id=index, minutes_ago=index + 1)

    book = workbook(audit_client.get(URL))

    assert len(cells(book["Audit trail"])) - 1 == 2
    text = cover_text(book)
    assert "INCOMPLETE" in text
    assert "Entries in this file 2" in text


def test_an_uncapped_file_does_not_cry_wolf(audit_client):
    an_entry()

    assert "INCOMPLETE" not in cover_text(workbook(audit_client.get(URL)))


# --------------------------------------------------------------------------------------
# The gate
# --------------------------------------------------------------------------------------
def test_the_export_is_behind_the_same_section_as_the_screen(client_for, db):
    """
    Who may take the record of who opened a farmer's Aadhaar card out of the building is the
    same question as who may read it on screen — and a narrower gate on the screen than on the
    file would be no gate at all.
    """
    user = User.objects.create_user(
        username="clerk",
        password="pw-for-tests-only",
        full_name="Rate clerk",
        role=Role.ADMIN,
        portal_sections=[PortalSection.RATES, PortalSection.REPORTS],
    )

    assert client_for(user).get(URL).status_code == 403


def test_the_export_is_read_only(audit_client):
    """An audit trail with a write path is not one, and that includes its export."""
    assert audit_client.post(URL, {}, format="json").status_code == 405
    assert audit_client.delete(URL).status_code == 405


def test_the_response_is_a_named_workbook(audit_client):
    an_entry()

    response = audit_client.get(URL)

    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    stamp = timezone.localdate().isoformat()
    assert response["Content-Disposition"] == f'attachment; filename="audit-trail-{stamp}.xlsx"'
