"""
The audit trail, as a workbook somebody can hand to an auditor (W19).

The screen reads the trail. This is for the other thing that happens to it: an auditor asks
for the record, and the answer cannot be "log in and scroll". They want a file — one that can
be filtered, sorted, pivoted, attached to a finding, and read in a year by somebody who has
never seen this product. That file has to carry its own provenance, because a spreadsheet of
rows with no statement of where it came from proves nothing at all.

**Four sheets, and only one of them is the data.**

``About this export`` opens first, and is the reason this is a workbook rather than a CSV. It
says who took the file, when, with which filters, how many rows that came to, and the request
id of the entry this very export wrote to the trail on its way out. An auditor holding the
file can go back to the running system, find that row, and confirm the file was produced when
it says it was. A CSV cannot say any of that without polluting its own columns.

``Audit trail`` is every row the screen's filters admit, in the screen's own order.

``Personal data`` is the ``pii_access`` rows on their own. They are already in the trail sheet
and this repeats them deliberately: SRS §16 asks for a record of who read personal data, and
"it is in there, filter for it" is not an answer to that question. It is the sheet most of
these files will be opened for.

``Who did what`` counts the file by person — entries each, personal-data reads each, first and
last activity. One person doing forty things is one person to ask about it, and that is the
shape an auditor works in.

**Built in memory, and capped.** An xlsx is a zip and cannot be valid until its central
directory is written, so there is nothing to stream: the file has to be finished before its
first byte is honest. The trail, unlike every other export here, grows without bound — so
``MAX_ROWS`` is a real ceiling rather than a tripwire, and when it bites the cover sheet says
so in red. A truncated file that does not admit it is evidence of the wrong thing.

**The rows are built by `serialise`**, the same function the screen renders from. The sentence
in the workbook is the sentence on the screen, down to the wording, because an auditor reading
"Completed AI event 64" while an operator reads something else about the same row is exactly
the confusion this file exists to prevent.
"""

from __future__ import annotations

from django.http import HttpResponse
from django.utils import timezone
from drf_spectacular.utils import OpenApiParameter, extend_schema
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view, permission_classes

from apps.accounts.models import PortalSection
from apps.core.audit_api import ACTIONS, entity_label, filtered, serialise
from apps.core.models import AuditLog
from apps.core.permissions import IsAdmin, in_section
from apps.core.services import record_audit

#: A hard ceiling, not a tripwire. Every other export in this codebase is bounded by a
#: population — farmers, Maits, a month of events. The trail is bounded by nothing: it grows
#: for as long as the platform runs. Somebody asking for "everything" on a three-year-old
#: installation would otherwise ask this process to hold several million rows in memory.
MAX_ROWS = 50_000

# The portal's ink, so the file looks like it came from the same product as the screen.
HEADER_FILL = PatternFill("solid", fgColor="253D4E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
NOTE_FONT = Font(color="475467", size=9)
#: Not a quiet grey note. A line saying the file is incomplete, or that it holds personal
#: data, has to survive being skimmed past on the way to the sheet tabs.
WARN_FONT = Font(color="B42318", bold=True)

#: The one action this file is usually opened for, tinted the way the screen tints it. An
#: auditor scrolling the trail sheet should not have to read the Action column to find them.
PII_FILL = PatternFill("solid", fgColor="FDECEC")

#: ``(header, width, force_text)``. Anything that is an identifier is text: a record id like
#: ``001302`` loses its leading zero as a number, an IP address becomes a date in some
#: locales, and a request id in scientific notation is not a request id any more.
COLUMNS = [
    ("Date", 12, True),
    ("Time", 10, True),
    ("Action", 20, False),
    ("What happened", 46, False),
    ("Record type", 16, False),
    ("Record", 14, True),
    ("Who", 24, False),
    ("Username", 18, True),
    ("Role", 18, False),
    ("IP address", 16, True),
    ("Request id", 34, True),
    ("Changed", 44, False),
    ("Details", 44, False),
]

PEOPLE_COLUMNS = [
    ("Who", 24, False),
    ("Username", 18, True),
    ("Role", 18, False),
    ("Entries", 10, False),
    ("Personal data reads", 20, False),
    ("First in this file", 20, True),
    ("Last in this file", 20, True),
]


def _at(title: str) -> int:
    """A column's index, by name. Inserting a column should not silently move a lookup."""
    return next(i for i, (name, _width, _text) in enumerate(COLUMNS) if name == title)


def _changed_text(changes: list[dict]) -> str:
    """
    A before/after pair as one cell.

    An arrow rather than two columns: a spreadsheet cannot hold a variable number of changed
    fields per row without either exploding the column count or losing which value belonged
    to which field.
    """
    return "; ".join(
        f"{change['field']}: {change['from'] or '—'} → {change['to'] or '—'}" for change in changes
    )


def _details_text(facts: list[dict]) -> str:
    return "; ".join(f"{fact['label']}: {fact['value']}" for fact in facts)


def _row(entry: AuditLog, data: dict) -> list:
    # Local time, split into two columns. One column of ISO timestamps cannot be grouped by
    # day in a pivot table without a formula, and "which days was this person working" is a
    # question every auditor asks.
    moment = timezone.localtime(entry.created_at)
    actor = data["actor"]
    return [
        moment.date().isoformat(),
        moment.strftime("%H:%M:%S"),
        data["action_label"],
        data["summary"],
        data["entity_label"],
        data["entity_id"],
        actor["name"],
        actor["username"],
        actor["role"],
        data["ip_address"],
        data["request_id"],
        _changed_text(data["changes"]),
        _details_text(data["facts"]),
    ]


def describe_filters(params) -> list[tuple[str, str]]:
    """
    The filters, in words rather than as a query string.

    ``action=pii_access&date_from=2026-08-01`` is what was sent; "Action: Personal data read"
    and "From 2026-08-01" is what somebody reading the file in a year needs, because the scope
    of the evidence is the first thing they have to establish.
    """
    described: list[tuple[str, str]] = []

    term = (params.get("search") or "").strip()
    if term:
        described.append(("Search", term))

    action = params.get("action")
    if action:
        described.append(("Action", ACTIONS.get(action, (action, None))[0]))

    entity_type = params.get("entity_type")
    if entity_type:
        described.append(("Record type", entity_label(entity_type)))

    actor = params.get("actor")
    if actor:
        described.append(("Person", actor))

    date_from, date_to = params.get("date_from"), params.get("date_to")
    if date_from and date_to:
        described.append(("Dates", f"{date_from} to {date_to}, inclusive"))
    elif date_from:
        described.append(("Dates", f"{date_from} onwards"))
    elif date_to:
        described.append(("Dates", f"up to and including {date_to}"))

    return described


def _write_header(sheet, columns, *, row: int) -> None:
    for index, (title, width, _force_text) in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_row(sheet, columns, values, *, line: int, tint) -> int:
    for index, value in enumerate(values, start=1):
        cell = sheet.cell(row=line, column=index, value=value)
        if columns[index - 1][2]:
            cell.number_format = "@"
        if tint is not None:
            cell.fill = tint
    return line + 1


def _finish(sheet, columns, *, header_row: int, next_row: int) -> None:
    """Freeze the header and let the sheet be filtered — sorting a list is how it gets read."""
    sheet.freeze_panes = f"A{header_row + 1}"
    if next_row > header_row + 1:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{next_row - 1}"


def _cover(sheet, *, actor, exported_at, filters, rows, truncated, request_id) -> None:
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 84

    line = 1

    def pair(label, value, *, font=None) -> None:
        nonlocal line
        left = sheet.cell(row=line, column=1, value=label)
        left.font = LABEL_FONT
        right = sheet.cell(row=line, column=2, value=value)
        right.alignment = Alignment(wrap_text=True, vertical="top")
        if font is not None:
            right.font = font
        line += 1

    def note(text, *, font=NOTE_FONT) -> None:
        nonlocal line
        cell = sheet.cell(row=line, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        line += 1

    title = sheet.cell(row=line, column=1, value="Mait AI — audit trail")
    title.font = TITLE_FONT
    line += 2

    # Who took it and when, first. This is the provenance an auditor establishes before
    # reading a single row, and a file that makes them hunt for it is a file they distrust.
    if actor is not None:
        name = actor.full_name or actor.username
        pair("Exported by", f"{name} ({actor.username}, {actor.get_role_display()})")
    else:
        pair("Exported by", "Unknown")
    pair("Exported at", exported_at.strftime("%d %b %Y, %H:%M:%S %Z"))
    pair("Entries in this file", rows)
    line += 1

    if filters:
        pair("Filters applied", "")
        for label, value in filters:
            pair(f"    {label}", value)
    else:
        pair("Filters applied", "None — the whole trail, newest first")
    line += 1

    # The thread back to the running system. Somebody holding this file can find this exact
    # request in the live trail and confirm the file was produced when it claims to have
    # been — which is the difference between a spreadsheet and evidence.
    if request_id:
        pair("Recorded in the trail as", request_id)
        note(
            "    This export wrote its own entry to the trail before building. Search the "
            "running system for that request id to confirm when this file was produced, and "
            "by whom."
        )
        line += 1

    if truncated:
        pair(
            "INCOMPLETE",
            f"This file was capped at {MAX_ROWS:,} entries. The filters matched more than "
            "that, and the oldest matching entries are not here. Narrow the date range and "
            "export again.",
            font=WARN_FONT,
        )
        line += 1

    note(
        "CONTAINS PERSONAL DATA — this trail records who opened farmers' identity documents "
        "and who exported files of personal details. Handle the file accordingly.",
        font=WARN_FONT,
    )
    line += 1

    note("Audit trail  ·  every entry the filters above admit, newest first.")
    note(
        "Personal data  ·  the same rows narrowed to reads of personal data (SRS §16). They "
        "appear in both sheets on purpose."
    )
    note("Who did what  ·  this file counted by person, busiest first.")
    line += 1

    note(
        "The trail is append-only. Nothing in this platform can edit or delete an entry: "
        "there is no write path to the table outside the one function that records to it, and "
        "the endpoint this file came from answers only GET. Entries are written at the moment "
        "of the action, not reconstructed afterwards."
    )
    note(
        "Times are Asia/Kolkata. Identifiers — records, usernames, IP addresses, request "
        "ids — are stored as text so that leading zeros and long digit strings survive the "
        "trip into a spreadsheet."
    )


def build_audit_workbook(queryset, *, actor=None, filters=None, request_id: str = "") -> Workbook:
    """
    The whole file, from a queryset the caller has already narrowed.

    The order is the caller's, and it is the screen's: this is taken from a screen and it
    should match the screen it was taken from, down to the sort.
    """
    exported_at = timezone.localtime()

    workbook = Workbook()
    cover = workbook.active
    cover.title = "About this export"
    trail = workbook.create_sheet("Audit trail")
    personal = workbook.create_sheet("Personal data")
    people = workbook.create_sheet("Who did what")

    _write_header(trail, COLUMNS, row=1)
    _write_header(personal, COLUMNS, row=1)
    _write_header(people, PEOPLE_COLUMNS, row=1)

    date_at, time_at = _at("Date"), _at("Time")

    # Counted while the rows are written, not queried separately. Two passes over a table that
    # is still being appended to can disagree, and a summary sheet that does not add up to the
    # sheet beside it is worse than no summary sheet — it makes the reader doubt both.
    tally: dict[str, dict] = {}

    line, pii_line, count, truncated = 2, 2, 0, False
    for entry in queryset.iterator(chunk_size=500):
        if count >= MAX_ROWS:
            truncated = True
            break

        data = serialise(entry)
        values = _row(entry, data)
        is_pii = entry.action == AuditLog.Action.PII_ACCESS

        line = _write_row(trail, COLUMNS, values, line=line, tint=PII_FILL if is_pii else None)
        if is_pii:
            pii_line = _write_row(personal, COLUMNS, values, line=pii_line, tint=PII_FILL)

        who = data["actor"]
        seen = tally.setdefault(
            who["username"] or who["name"],
            {
                "name": who["name"],
                "username": who["username"],
                "role": who["role"],
                "entries": 0,
                "pii": 0,
                "first": None,
                "last": None,
            },
        )
        seen["entries"] += 1
        seen["pii"] += 1 if is_pii else 0
        stamp = f"{values[date_at]} {values[time_at]}"
        # Compared rather than assumed. The queryset arrives newest first, so the first row
        # seen for somebody is their most recent activity — but this builder is reachable from
        # a shell with any ordering, and a first/last pair that silently inverts is a lie.
        seen["first"] = stamp if seen["first"] is None else min(seen["first"], stamp)
        seen["last"] = stamp if seen["last"] is None else max(seen["last"], stamp)
        count += 1

    _finish(trail, COLUMNS, header_row=1, next_row=line)
    _finish(personal, COLUMNS, header_row=1, next_row=pii_line)

    if pii_line == 2:
        # Said, rather than left blank. An empty sheet reads as one that failed to load; this
        # one is a finding in its own right.
        cell = personal.cell(
            row=2, column=1, value="No personal data was read in the entries this file covers."
        )
        cell.font = NOTE_FONT

    # Busiest first, and whoever read personal data ahead of whoever did not: an auditor works
    # down from the top of this sheet.
    ordered = sorted(tally.values(), key=lambda item: (-item["pii"], -item["entries"]))
    people_line = 2
    for item in ordered:
        people_line = _write_row(
            people,
            PEOPLE_COLUMNS,
            [
                item["name"],
                item["username"],
                item["role"],
                item["entries"],
                item["pii"],
                item["first"] or "",
                item["last"] or "",
            ],
            line=people_line,
            tint=PII_FILL if item["pii"] else None,
        )
    _finish(people, PEOPLE_COLUMNS, header_row=1, next_row=people_line)

    _cover(
        cover,
        actor=actor,
        exported_at=exported_at,
        filters=filters or [],
        rows=count,
        truncated=truncated,
        request_id=request_id,
    )

    # Opens on the cover, whatever the reader's Excel last remembered about this file.
    workbook.active = 0
    return workbook


def audit_workbook_response(
    queryset, *, actor=None, filters=None, request_id: str = ""
) -> HttpResponse:
    workbook = build_audit_workbook(queryset, actor=actor, filters=filters, request_id=request_id)
    stamp = timezone.localdate().isoformat()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="audit-trail-{stamp}.xlsx"'
    workbook.save(response)
    workbook.close()
    return response


@extend_schema(
    tags=["core"],
    summary="Take the audit trail away as a workbook",
    description=(
        "The trail as an xlsx an auditor can be handed, narrowed by the same filters the "
        "screen uses — `search`, `action`, `entity_type`, `actor`, `date_from`, `date_to`. "
        "`limit` and `offset` are not read: this is the query, not the page of it somebody "
        "happened to be looking at.\n\n"
        "Four sheets. **About this export** carries the provenance — who took it, when, with "
        "which filters, and the request id of the entry this export wrote to the trail on its "
        "way out, so the file can be tied back to the running system. **Audit trail** is the "
        "rows. **Personal data** repeats the `pii_access` rows on their own, because SRS §16 "
        'asks for a record of who read personal data and "filter for it" is not an answer. '
        "**Who did what** counts the file by person.\n\n"
        "Capped at 50,000 entries, oldest dropped first; when the cap bites the cover sheet "
        "says so in red rather than handing back a short file that looks complete.\n\n"
        "**Taking this file is itself a `pii_access` entry** in the trail, with the filters it "
        "ran. A record of who reads the record is the whole point of the record."
    ),
    parameters=[
        OpenApiParameter("search", description="Record, person, request id or IP", type=str),
        OpenApiParameter("action", description="One of the AuditLog actions", type=str),
        OpenApiParameter("entity_type", description="Record type", type=str),
        OpenApiParameter("actor", description="Username", type=str),
        OpenApiParameter("date_from", description="YYYY-MM-DD, inclusive", type=str),
        OpenApiParameter("date_to", description="YYYY-MM-DD, inclusive", type=str),
    ],
    responses={200: bytes},
)
@api_view(["GET"])
@permission_classes([IsAdmin, in_section(PortalSection.LOGS)])
def audit_trail_export(request):
    """
    Build the workbook, and record that it was built.

    The audit row is written *before* the file, deliberately and in that order. An export that
    fell over halfway still means somebody asked for the trail and the server started handing
    it over; a trail that only records the exports which completed is a trail with a gap
    exactly where somebody would want one.

    That row is then kept out of the file it describes. The trail is evaluated after the row
    exists, so without the exclusion every export would open on the record of its own making —
    an auditor reading that they themselves asked this question, nought seconds ago, learns
    nothing from it, and it would inflate their own line on the Who did what sheet. Worse, it
    would mean the Personal data sheet could never be empty, and "nobody read personal data in
    this window" is a finding worth being able to state.

    *Previous* exports stay in, and belong there: who has taken this trail out of the building
    before is exactly the kind of thing the file is read to establish. It is only the export in
    hand that is not news. The cover carries its request id, so the row can still be found in
    the running system.
    """
    entry = record_audit(
        action=AuditLog.Action.PII_ACCESS,
        entity_type="report",
        # Reads as "Exported audit trail" once `describe` has had it. The screen and the file
        # both go through that function, so this string is user-facing prose, not a key.
        entity_id="audit_trail",
        request=request,
        meta={
            "filters": {
                key: value
                for key, value in request.query_params.items()
                if key not in ("limit", "offset")
            }
        },
    )

    return audit_workbook_response(
        # Through `filtered`, the screen's own definition — see the note on that function.
        # `limit` and `offset` are not passed and are not read: they are how the table pages,
        # and a file built from them would hand back the fifty rows that happened to be on
        # screen. That file looks complete, is not, and says nothing about it.
        #
        # Minus this export's own entry — see the docstring. By primary key rather than by
        # shape: two people exporting in the same second must each drop their own row and
        # keep the other's.
        filtered(request.query_params).exclude(pk=entry.pk),
        actor=entry.actor,
        filters=describe_filters(request.query_params),
        request_id=entry.request_id or "",
    )
