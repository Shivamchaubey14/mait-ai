"""
The Mait payout month as the workbook the office already knows (W18).

Deliberately the same shape as the sheet this replaced, down to the column order and the two
tabs — because the people who read it have read it every month for years, and a report that
is *better* but unfamiliar gets checked against the old one by hand, which is the work this
was meant to remove. The improvements are underneath: the numbers come from the event log
rather than a keyboard, and the rate legend at the foot is what the file was actually built
with rather than a note somebody remembered to update.

**xlsx, not CSV**, for the reason ``masterdata.exports`` sets out at length: this file is
mostly identifiers, and Excel eats them coming in from a CSV. A fifteen-digit account number
becomes ``7.52210E+14``, an IFSC survives but a vendor code loses its leading zeros, and the
number that arrives rounded is the number somebody is about to pay money into. Every code
cell is written as text.

**It carries full bank details, and that is the point of it.** This is a payment instruction:
a masked account number cannot be paid into, so the exception the non-member roster makes is
made here too, and held in place the same way — Admin only, behind the Mait payment section,
and audit-logged as ``pii_access`` with the month it covered. The preview on the screen masks
the account and the PAN, because a screen is read over a shoulder and a file is not.

**The disclosure is in the workbook's properties, not in a row above the data.** It began as a
banner on row 1 and that was the wrong place for it: the first thing anybody does with this
file is sort or filter it, and a stray row above the headings makes Excel guess the wrong
header row every single time. So the sheet starts with its column names, the way the office's
own does, and what the banner said now lives in the document properties, on the screen before
the download, and on the audit log.

**Built in memory, not streamed.** An xlsx is a zip and is not valid until its central
directory is written, so there is nothing to stream; and the row count is the number of
working Maits, which is people rather than events.
"""

from __future__ import annotations

from datetime import date

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.payments.payout import build_payout

# The portal's ink and washes, so the file looks like it came from the same product as the
# assignment sheet and the non-member roster.
HEADER_FILL = PatternFill("solid", fgColor="253D4E")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="EAF7F1")
TOTAL_FONT = Font(bold=True)
LEGEND_FILL = PatternFill("solid", fgColor="FFF8E9")
LEGEND_FONT = Font(bold=True, size=10)
#: A payout that came out below zero — more issued than earned. Tinted rather than filtered
#: out: it is the row somebody has to look at before the file goes to the bank.
NEGATIVE_FILL = PatternFill("solid", fgColor="FDECEA")
THIN = Side(style="thin", color="D4DBE0")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = "#,##0"
#: Litres come off the ledger whole, but the format allows a half so a future fractional
#: issue is not silently rounded on the way into the file.
QUANTITY = "#,##0.##"

#: ``(header, width, kind)``. ``text`` cells are identifiers Excel would otherwise mangle;
#: ``money`` and ``qty`` stay genuinely numeric so the office can pivot and re-total them.
COLUMNS = [
    ("S.No.", 7, "plain"),
    ("MCC NAME", 20, "plain"),
    ("MAITS NAME", 26, "plain"),
    ("AI PERFORMED", 13, "qty"),
    ("TOTAL AMOUNT", 13, "money"),
    ("FIXED AMOUNT", 13, "money"),
    ("TOTAL", 12, "money"),
    ("SEMEN", 9, "qty"),
    ("LN2", 9, "qty"),
    ("SHEETH", 9, "qty"),
    ("GLOVES", 9, "qty"),
    ("TOTAL DEDUCTION", 15, "money"),
    ("AFTER DEDUCTION", 15, "money"),
    ("TAGGING", 10, "money"),
    ("NET PAYABLE", 13, "money"),
    ("Account No", 22, "text"),
    ("IFSC", 14, "text"),
    ("PANCARD", 14, "text"),
    ("VENDOR", 14, "text"),
]

#: Which material column each quantity heading is fed by, in the sheet's own order.
QUANTITY_KEYS = ["semen", "ln2", "sheath", "gloves"]

MONTH_NAMES = [
    "JAN",
    "FEB",
    "MAR",
    "APR",
    "MAY",
    "JUN",
    "JUL",
    "AUG",
    "SEP",
    "OCT",
    "NOV",
    "DEC",
]


def _month_label(month: date) -> str:
    """``MAR 2026`` — the way the office names the file it is looking for."""
    return f"{MONTH_NAMES[month.month - 1]} {month.year}"


def _write(sheet, row, column, value, *, kind="plain", fill=None, font=None):
    cell = sheet.cell(row=row, column=column, value=value)
    if kind == "text":
        cell.number_format = "@"
    elif kind == "money":
        cell.number_format = MONEY
    elif kind == "qty":
        cell.number_format = QUANTITY
    cell.border = GRID
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    return cell


def _payout_sheet(sheet, report: dict) -> None:
    scheme = report["scheme"]
    rates = report["rates"]

    # The column names are the first row, as they are on the office's own sheet. A banner
    # above them used to say what the file holds; it was in the way — the first thing anybody
    # does with this workbook is sort or filter it, and a stray row above the headings makes
    # Excel guess the wrong header row every time. What it said has not been dropped, only
    # moved somewhere that does not sit in the data: the screen warns before the download, the
    # workbook's own properties carry it (see `build_payout_workbook`), and every export is on
    # the audit log with the month it covered.
    for index, (title, width, _kind) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = GRID
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width

    line = 2
    for serial, row in enumerate(report["rows"], start=1):
        tint = NEGATIVE_FILL if row.is_overdrawn else None
        values = [
            serial,
            row.mcc_name,
            row.mait_name,
            row.ai_performed,
            float(row.commission),
            float(row.fixed_amount),
            float(row.gross),
            *(row.quantities.get(key, 0) for key in QUANTITY_KEYS),
            float(row.deduction),
            float(row.after_deduction),
            float(row.tagging),
            float(row.net_payable),
            # In full, all three. This file is a payment instruction — see the note at the top
            # of the module about why that is the exception here.
            row.bank_account_no,
            row.ifsc_code,
            row.pan_no,
            row.vendor_code,
        ]
        for index, value in enumerate(values, start=1):
            _write(sheet, line, index, value, kind=COLUMNS[index - 1][2], fill=tint)
        line += 1

    totals = report["totals"]
    _write(sheet, line, 3, "Total", fill=TOTAL_FILL, font=TOTAL_FONT)
    for index in (1, 2):
        _write(sheet, line, index, None, fill=TOTAL_FILL)
    total_values = [
        (4, totals["ai_performed"], "qty"),
        (5, float(totals["commission"]), "money"),
        (6, float(totals["fixed_amount"]), "money"),
        (7, float(totals["gross"]), "money"),
        *(
            (8 + offset, totals["quantities"].get(key, 0), "qty")
            for offset, key in enumerate(QUANTITY_KEYS)
        ),
        (12, float(totals["deduction"]), "money"),
        (13, float(totals["after_deduction"]), "money"),
        (14, float(totals["tagging"]), "money"),
        (15, float(totals["net_payable"]), "money"),
    ]
    for index, value, kind in total_values:
        _write(sheet, line, index, value, kind=kind, fill=TOTAL_FILL, font=TOTAL_FONT)
    # The identifier columns have no total, but the band should run the width of the table
    # rather than stop halfway across it.
    for index in range(16, len(COLUMNS) + 1):
        _write(sheet, line, index, None, fill=TOTAL_FILL)

    _legend(sheet, line + 3, scheme, rates)

    # Below the headings and past the three identity columns, so scrolling either way keeps
    # both the column names and whose row it is in view.
    sheet.freeze_panes = "D2"
    if report["rows"]:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{line - 1}"


def _legend(sheet, line: int, scheme, rates: dict) -> None:
    """
    The terms the numbers above were produced with, written as figures rather than as prose.

    The office sheet has always carried this block and it has always been typed by hand, which
    is exactly how a file comes to state one commission rate and be calculated on another.
    Here it is read from the same scheme object the rows were, so the two cannot disagree.
    """
    _write(sheet, line, 3, "AI Commission", fill=LEGEND_FILL, font=LEGEND_FONT)
    _write(
        sheet,
        line,
        4,
        f"Rs {scheme.commission_per_ai:.0f} Per AI",
        fill=LEGEND_FILL,
    )
    _write(sheet, line, 6, "MATERIAL", fill=LEGEND_FILL, font=LEGEND_FONT)
    _write(sheet, line, 7, "RATE PER PIECE", fill=LEGEND_FILL, font=LEGEND_FONT)

    if scheme.fixed_min_ai:
        _write(
            sheet,
            line + 1,
            3,
            f"At least {scheme.fixed_min_ai} AI in month",
            fill=LEGEND_FILL,
        )
        _write(
            sheet,
            line + 1,
            4,
            f"Rs {scheme.monthly_fixed_amount:.0f} (Fixed)",
            fill=LEGEND_FILL,
        )

    materials = [("SEMEN", rates.get("semen")), ("SHEETH", rates.get("sheath"))]
    materials += [("GLOVES", rates.get("gloves")), ("LN2", rates.get("ln2"))]
    for offset, (label, rate) in enumerate(materials, start=1):
        _write(sheet, line + offset, 6, label, fill=LEGEND_FILL)
        _write(sheet, line + offset, 7, float(rate or 0), kind="money", fill=LEGEND_FILL)


def _deduction_sheet(sheet, report: dict) -> None:
    """
    The office sheet's second tab: how many inseminations finance recovers, per MCC.

    Kept as its own tab rather than folded into the payout, because it answers a different
    question for a different desk — the payout is paid *to* Maits, this is recovered *from*
    members — and the two are counted on different keys. Collapsing them would invite somebody
    to reconcile one against the other and find a difference that is not an error.
    """
    month = report["month"]
    title = sheet.cell(row=2, column=3, value=f"{_month_label(month)} AI DEDUCTION")
    title.font = Font(bold=True, size=11)

    for offset, label in enumerate(("MCC Name", "DEDUCTION AI")):
        cell = sheet.cell(row=3, column=3 + offset, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = GRID
    sheet.column_dimensions["C"].width = 26
    sheet.column_dimensions["D"].width = 16

    line = 4
    for entry in report["deductions"]:
        _write(sheet, line, 3, entry["mcc_name"])
        _write(sheet, line, 4, entry["ai_count"], kind="qty")
        line += 1

    _write(sheet, line, 3, "TOTAL", fill=TOTAL_FILL, font=TOTAL_FONT)
    _write(
        sheet,
        line,
        4,
        sum(entry["ai_count"] for entry in report["deductions"]),
        kind="qty",
        fill=TOTAL_FILL,
        font=TOTAL_FONT,
    )

    note = sheet.cell(
        row=line + 3,
        column=3,
        value=(
            "Counted where the insemination happened, not where the Mait is posted — this "
            "tab is settled by collection point. Members whose charge is deducted from a "
            "milk payment only; cash and online payments are already settled."
        ),
    )
    note.font = Font(size=9, color="7A8893")
    note.alignment = Alignment(wrap_text=True, vertical="top")


def build_payout_workbook(year: int, month: int) -> tuple[Workbook, dict]:
    """The two-tab workbook, and the report it was built from so a caller can audit it."""
    report = build_payout(year, month)

    workbook = Workbook()
    payout = workbook.active
    payout.title = f"{MONTH_NAMES[report['month'].month - 1]}-PAYMENT"
    _payout_sheet(payout, report)

    _deduction_sheet(workbook.create_sheet("DEDUCTION"), report)

    # What the file is and what it holds, in the workbook's properties rather than in a row
    # above the data. Excel shows these under File → Info, they travel with the file, and
    # unlike a banner row they are not something a sort or a filter has to be told to skip.
    workbook.properties.title = f"Mait payment — {_month_label(report['month'])}"
    workbook.properties.subject = (
        "Contains full bank account numbers, IFSC codes and PANs — handle as personal data"
    )
    workbook.properties.creator = "Mait AI"
    workbook.properties.created = timezone.now().replace(tzinfo=None)
    return workbook, report


def payout_workbook_response(year: int, month: int) -> tuple[HttpResponse, dict]:
    workbook, report = build_payout_workbook(year, month)
    stamp = f"{report['month'].year}-{report['month'].month:02d}"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="mait-payment-{stamp}.xlsx"'
    workbook.save(response)
    workbook.close()
    return response, report
